"""Owner-authenticated admin API, on in-memory SQLite (no Supabase).

Covers the new auth surface (dev token + Supabase token introspection + email
allowlist, and that the dev token is ignored in production), guest/question CRUD,
tenant scoping, the responses/summary rollups, and CSV export.

The Supabase path is exercised by monkeypatching `app.auth.verify_supabase_token`
(token introspection) so the tests stay fully offline.
"""
from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.auth as auth_module
from app.auth import get_current_owner
from app.config import Settings, get_settings
from app.db import Base, get_db
from app.main import app
from app.models import (
    Answer,
    Companion,
    CompanionKind,
    Guest,
    InviteTier,
    Question,
    QuestionApplies,
    QuestionScope,
    QuestionType,
    Rsvp,
    Wedding,
)

DEV_TOKEN = "dev-secret-token"
ADMIN_EMAIL = "owner@example.com"


def _settings(**overrides) -> Settings:
    base = dict(
        environment="development",
        dev_admin_token=DEV_TOKEN,
        supabase_url="https://example.supabase.co",
        supabase_publishable_key="sb_publishable_test",
        admin_emails=ADMIN_EMAIL,
    )
    base.update(overrides)
    # _env_file=None keeps tests deterministic (ignore any .env/.env.local).
    return Settings(_env_file=None, **base)


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    try:
        yield db
    finally:
        db.close()
        Base.metadata.drop_all(engine)


@pytest.fixture
def make_client(db_session):
    """Factory: build a TestClient with given settings overrides."""

    def _make(**setting_overrides) -> TestClient:
        s = _settings(**setting_overrides)
        app.dependency_overrides[get_db] = lambda: db_session
        app.dependency_overrides[get_settings] = lambda: s
        return TestClient(app)

    yield _make
    app.dependency_overrides.clear()


@pytest.fixture
def client(make_client):
    return make_client()


def dev_auth() -> dict:
    return {"Authorization": f"Bearer {DEV_TOKEN}"}


def supabase_auth() -> dict:
    """A bearer token that the patched introspection will accept (any value)."""
    return {"Authorization": "Bearer supabase-access-token"}


def patch_introspection(monkeypatch, *, email: str | None = ADMIN_EMAIL, valid: bool = True):
    """Stand in for Supabase's /auth/v1/user. `valid=False` simulates a rejected
    token (401); otherwise return a user with the given email."""

    def fake(settings, token):
        if not valid:
            from app.auth import _unauthorized

            raise _unauthorized("Invalid or expired session")
        return {"id": "user-1", "email": email}

    monkeypatch.setattr(auth_module, "verify_supabase_token", fake)


@pytest.fixture
def wedding(db_session):
    w = Wedding(
        slug="alex-and-sam",
        couple_names="Alex & Sam",
        status="active",
        event_details={"venue": "The Garden Hall"},
        content={},
        theme_tokens=None,
    )
    db_session.add(w)
    db_session.commit()
    return w


def _add_guest(db, wedding, *, slug, name, tier=InviteTier.solo, invited=True):
    g = Guest(
        wedding_id=wedding.id, slug=slug, name=name, invite_tier=tier, invited=invited,
        greeting_name=name.split(" ")[0] if name else "Guest",
    )
    db.add(g)
    db.commit()
    return g


# --- Auth ------------------------------------------------------------------
def test_no_token_is_401(client, wedding):
    assert client.get("/api/admin/me").status_code == 401


def test_bad_token_is_401(client, wedding, monkeypatch):
    # A non-dev token is introspected; simulate Supabase rejecting it.
    patch_introspection(monkeypatch, valid=False)
    r = client.get("/api/admin/me", headers={"Authorization": "Bearer nope"})
    assert r.status_code == 401


def test_dev_token_grants_access(client, wedding):
    r = client.get("/api/admin/me", headers=dev_auth())
    assert r.status_code == 200
    body = r.json()
    assert body["via"] == "dev"
    assert body["wedding_slug"] == "alex-and-sam"


def test_valid_supabase_token_allowlisted(client, wedding, monkeypatch):
    patch_introspection(monkeypatch, email=ADMIN_EMAIL)
    r = client.get("/api/admin/me", headers=supabase_auth())
    assert r.status_code == 200
    assert r.json()["via"] == "supabase"
    assert r.json()["email"] == ADMIN_EMAIL


def test_supabase_token_not_allowlisted_is_403(client, wedding, monkeypatch):
    patch_introspection(monkeypatch, email="stranger@example.com")
    r = client.get("/api/admin/me", headers=supabase_auth())
    assert r.status_code == 403


def test_supabase_token_rejected_is_401(client, wedding, monkeypatch):
    patch_introspection(monkeypatch, valid=False)
    r = client.get("/api/admin/me", headers=supabase_auth())
    assert r.status_code == 401


def test_dev_token_ignored_in_production(make_client, wedding, monkeypatch):
    client = make_client(environment="production")
    # In production the dev token isn't honoured → falls through to introspection,
    # which here rejects it → 401.
    patch_introspection(monkeypatch, valid=False)
    r = client.get("/api/admin/me", headers=dev_auth())
    assert r.status_code == 401
    # A real allowlisted Supabase session still works in production.
    patch_introspection(monkeypatch, email=ADMIN_EMAIL)
    assert client.get("/api/admin/me", headers=supabase_auth()).status_code == 200


# --- Guest CRUD ------------------------------------------------------------
def test_list_guests_with_rollup(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="a-1", name="Riley", tier=InviteTier.plus_one)
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True))
    db_session.commit()
    rows = client.get("/api/admin/guests", headers=dev_auth()).json()
    assert len(rows) == 1
    assert rows[0]["invite_tier"] == "plus_one"
    assert rows[0]["rsvp_status"] == "attending"
    assert rows[0]["party_size"] == 1
    assert rows[0]["invite_path"] == "/i/a-1"


def test_create_guest_generates_link(client, wedding):
    r = client.post(
        "/api/admin/guests",
        headers=dev_auth(),
        json={"name": "New Person", "greeting_name": "New Person", "invite_tier": "plus_family", "side": "Alex"},
    )
    assert r.status_code == 201
    body = r.json()
    assert body["invite_tier"] == "plus_family"
    # The slug is a non-descript random token (no name leaked into the link).
    assert "new-person" not in body["slug"]
    assert all(c.isalnum() or c in "-_" for c in body["slug"]) and len(body["slug"]) >= 22
    assert body["invite_path"] == f"/i/{body['slug']}"
    assert body["rsvp_status"] == "pending"


def test_create_guest_rejects_bad_tier(client, wedding):
    r = client.post(
        "/api/admin/guests", headers=dev_auth(), json={"name": "X", "invite_tier": "vip"}
    )
    assert r.status_code == 422


def test_create_guest_validates_and_normalizes_contacts(client, wedding):
    # Bad email rejected.
    bad = client.post(
        "/api/admin/guests",
        headers=dev_auth(),
        json={"name": "Bad", "greeting_name": "Bad", "email": "nope"},
    )
    assert bad.status_code == 422
    # Valid contacts: phone normalized to E.164, email domain lowercased.
    ok = client.post(
        "/api/admin/guests",
        headers=dev_auth(),
        json={"name": "Reachable", "greeting_name": "Reachable", "email": "Me@Example.COM", "phone": "91234567"},
    )
    assert ok.status_code == 201
    body = ok.json()
    assert body["email"] == "Me@example.com"
    assert body["phone"] == "+6591234567"


def test_expected_party_size_set_clear_and_admin_only(client, db_session, wedding):
    # Owner can record an estimate at create time.
    created = client.post(
        "/api/admin/guests",
        headers=dev_auth(),
        json={"name": "Estimator", "greeting_name": "Estimator", "invite_tier": "plus_family", "expected_party_size": 4},
    )
    assert created.status_code == 201
    body = created.json()
    assert body["expected_party_size"] == 4
    gid, slug = body["id"], body["slug"]

    # Omitting it defaults to null (no estimate yet).
    none_body = client.post(
        "/api/admin/guests", headers=dev_auth(),
        json={"name": "No Est", "greeting_name": "No Est"},
    ).json()
    assert none_body["expected_party_size"] is None

    # Update can change it, and null clears it.
    changed = client.patch(
        f"/api/admin/guests/{gid}", headers=dev_auth(), json={"expected_party_size": 2}
    )
    assert changed.json()["expected_party_size"] == 2
    cleared = client.patch(
        f"/api/admin/guests/{gid}", headers=dev_auth(), json={"expected_party_size": None}
    )
    assert cleared.json()["expected_party_size"] is None

    # Negative is rejected by the schema.
    bad = client.post(
        "/api/admin/guests", headers=dev_auth(),
        json={"name": "Neg", "greeting_name": "Neg", "expected_party_size": -1},
    )
    assert bad.status_code == 422

    # Admin-only: it must never reach the guest-facing invite payload.
    invite = client.get(f"/api/i/{slug}")
    assert invite.status_code == 200
    assert "expected_party_size" not in invite.text


def test_update_guest_tier(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="u-1", name="Up Grade", tier=InviteTier.solo)
    r = client.patch(
        f"/api/admin/guests/{g.id}",
        headers=dev_auth(),
        json={"invite_tier": "plus_one", "relationship": "Cousin"},
    )
    assert r.status_code == 200
    assert r.json()["invite_tier"] == "plus_one"
    assert r.json()["relationship"] == "Cousin"


def test_delete_guest(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="d-1", name="Gone")
    assert client.delete(f"/api/admin/guests/{g.id}", headers=dev_auth()).status_code == 204
    assert client.get("/api/admin/guests", headers=dev_auth()).json() == []


# --- Companions (the +1 / kids on an RSVP) ---------------------------------
def _party_with_companions(db, wedding):
    """A plus_family guest attending with one adult + one child companion, plus a
    per-person dietary question (everyone) and a children-only Age question. Leo has
    an age answer. Returns (guest, adult, child, diet_q, age_q)."""
    diet_q = Question(
        wedding_id=wedding.id, prompt="Any dietary needs?", qtype=QuestionType.multi_choice,
        options=["Halal", "Vegetarian"], scope=QuestionScope.person,
        applies_to=QuestionApplies.everyone,
    )
    age_q = Question(
        wedding_id=wedding.id, prompt="Age", qtype=QuestionType.number,
        scope=QuestionScope.person, applies_to=QuestionApplies.children,
    )
    db.add_all([diet_q, age_q])
    g = _add_guest(db, wedding, slug="fam-1", name="Hasaan", tier=InviteTier.plus_family)
    rsvp = Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True)
    db.add(rsvp)
    db.flush()
    adult = Companion(wedding_id=wedding.id, rsvp_id=rsvp.id, kind=CompanionKind.adult, name="May")
    child = Companion(wedding_id=wedding.id, rsvp_id=rsvp.id, kind=CompanionKind.child, name="Leo")
    db.add_all([adult, child])
    db.flush()
    db.add(Answer(wedding_id=wedding.id, rsvp_id=rsvp.id, question_id=age_q.id,
                  companion_id=child.id, value={"number": 6}))
    db.commit()
    return g, adult, child, diet_q, age_q


def test_companion_id_in_guest_rollup(client, db_session, wedding):
    g, adult, child, _, _ = _party_with_companions(db_session, wedding)
    rows = client.get("/api/admin/guests", headers=dev_auth()).json()
    comps = {c["name"]: c for c in rows[0]["companions"]}
    assert comps["May"]["id"] == str(adult.id)
    assert comps["Leo"]["id"] == str(child.id)
    # Leo's age rides on his per-person answers now.
    assert comps["Leo"]["answers"][0]["value"] == {"number": 6}


def test_update_companion(client, db_session, wedding):
    _, adult, child, diet_q, age_q = _party_with_companions(db_session, wedding)
    # Edit the adult's name + replace their answers (the dietary question applies).
    r = client.patch(
        f"/api/admin/companions/{adult.id}",
        headers=dev_auth(),
        json={"name": "May Tan", "answers": [
            {"question_id": str(diet_q.id), "value": {"choices": ["Vegetarian"]}},
        ]},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["id"] == str(adult.id) and body["name"] == "May Tan"
    assert body["answers"][0]["value"] == {"choices": ["Vegetarian"]}
    # Edit the child's age answer.
    r2 = client.patch(
        f"/api/admin/companions/{child.id}", headers=dev_auth(),
        json={"answers": [{"question_id": str(age_q.id), "value": {"number": 7}}]},
    )
    assert r2.status_code == 200 and r2.json()["answers"][0]["value"] == {"number": 7}


def test_update_companion_rejects_inapplicable_answer(client, db_session, wedding):
    """Answering the children-only Age question for an adult companion → 422."""
    _, adult, _, _, age_q = _party_with_companions(db_session, wedding)
    r = client.patch(
        f"/api/admin/companions/{adult.id}", headers=dev_auth(),
        json={"answers": [{"question_id": str(age_q.id), "value": {"number": 40}}]},
    )
    assert r.status_code == 422


def test_delete_companion(client, db_session, wedding):
    g, adult, _, _, _ = _party_with_companions(db_session, wedding)
    assert client.delete(f"/api/admin/companions/{adult.id}", headers=dev_auth()).status_code == 204
    rows = client.get("/api/admin/guests", headers=dev_auth()).json()
    names = [c["name"] for c in rows[0]["companions"]]
    assert names == ["Leo"] and rows[0]["party_size"] == 2  # primary + remaining child


def test_companion_endpoints_are_tenant_scoped(make_client, db_session):
    a = Wedding(slug="wed-a", couple_names="A", status="active", owner_id="dev",
                event_details={}, content={})
    b = Wedding(slug="wed-b", couple_names="B", status="active", owner_id="other",
                event_details={}, content={})
    db_session.add_all([a, b])
    db_session.commit()
    _, _, child_b, _, _ = _party_with_companions(db_session, b)
    client = make_client()  # dev owner → wedding A
    assert client.patch(
        f"/api/admin/companions/{child_b.id}", headers=dev_auth(), json={"name": "x"}
    ).status_code == 404
    assert client.delete(
        f"/api/admin/companions/{child_b.id}", headers=dev_auth()
    ).status_code == 404


# --- Tenant scoping --------------------------------------------------------
def test_admin_only_sees_own_wedding(make_client, db_session):
    # Two weddings; the dev owner is claimed onto wedding A.
    a = Wedding(slug="wed-a", couple_names="A", status="active", owner_id="dev",
                event_details={}, content={})
    b = Wedding(slug="wed-b", couple_names="B", status="active", owner_id="other",
                event_details={}, content={})
    db_session.add_all([a, b])
    db_session.commit()
    ga = _add_guest(db_session, a, slug="ga", name="Guest A")
    gb = _add_guest(db_session, b, slug="gb", name="Guest B")

    client = make_client()
    rows = client.get("/api/admin/guests", headers=dev_auth()).json()
    assert [r["name"] for r in rows] == ["Guest A"]
    # Cannot touch wedding B's guest.
    assert client.patch(
        f"/api/admin/guests/{gb.id}", headers=dev_auth(), json={"name": "Hijack"}
    ).status_code == 404
    assert client.delete(f"/api/admin/guests/{gb.id}", headers=dev_auth()).status_code == 404


# --- Questions CRUD --------------------------------------------------------
def test_question_crud(client, wedding):
    created = client.post(
        "/api/admin/questions",
        headers=dev_auth(),
        json={"prompt": "Song request?", "qtype": "text", "required": False},
    )
    assert created.status_code == 201
    qid = created.json()["id"]
    # Scope/applies_to default to invitee/everyone when omitted.
    assert created.json()["scope"] == "invitee"
    assert created.json()["applies_to"] == "everyone"

    listed = client.get("/api/admin/questions", headers=dev_auth()).json()
    assert len(listed) == 1 and listed[0]["prompt"] == "Song request?"

    # Patch type/scope/applies_to together (e.g. make it a per-person multi-select).
    patched = client.patch(
        f"/api/admin/questions/{qid}", headers=dev_auth(),
        json={"required": True, "qtype": "multi_choice", "options": ["A", "B"],
              "scope": "person", "applies_to": "children"},
    )
    assert patched.status_code == 200
    assert patched.json()["required"] is True
    assert patched.json()["qtype"] == "multi_choice"
    assert patched.json()["scope"] == "person"
    assert patched.json()["applies_to"] == "children"

    assert client.delete(f"/api/admin/questions/{qid}", headers=dev_auth()).status_code == 204
    assert client.get("/api/admin/questions", headers=dev_auth()).json() == []


def test_question_create_rejects_extra_field(client, wedding):
    r = client.post(
        "/api/admin/questions",
        headers=dev_auth(),
        json={"prompt": "Q", "qtype": "text", "wedding_id": "x"},
    )
    assert r.status_code == 422


# --- Responses + summary ---------------------------------------------------
def _seed_responses(db, wedding):
    # An invitee-scope choice question (party answer) + a children-only Age question.
    q = Question(wedding_id=wedding.id, prompt="Main course", qtype=QuestionType.choice,
                 options=["Fish", "Chicken"], required=True)
    age_q = Question(wedding_id=wedding.id, prompt="Age", qtype=QuestionType.number,
                     scope=QuestionScope.person, applies_to=QuestionApplies.children)
    db.add_all([q, age_q])
    db.commit()

    g1 = _add_guest(db, wedding, slug="att-1", name="Attendee One", tier=InviteTier.plus_family)
    r1 = Rsvp(wedding_id=wedding.id, guest_id=g1.id, attending=True)
    db.add(r1)
    db.flush()
    kid = Companion(wedding_id=wedding.id, rsvp_id=r1.id, kind=CompanionKind.child, name="Kiddo")
    r1.companions[:] = [
        Companion(wedding_id=wedding.id, rsvp_id=r1.id, kind=CompanionKind.adult, name="Plus One"),
        kid,
    ]
    db.flush()
    r1.answers[:] = [
        Answer(wedding_id=wedding.id, question_id=q.id, value={"choice": "Fish"}),
        Answer(wedding_id=wedding.id, question_id=age_q.id, companion_id=kid.id, value={"number": 4}),
    ]
    db.commit()

    g2 = _add_guest(db, wedding, slug="dec-1", name="Decliner")
    db.add(Rsvp(wedding_id=wedding.id, guest_id=g2.id, attending=False))
    # g3 pending (no rsvp)
    _add_guest(db, wedding, slug="pen-1", name="Pending")
    db.commit()
    return q


def test_summary_rollup(client, db_session, wedding):
    _seed_responses(db_session, wedding)
    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s["total_guests"] == 3
    assert s["attending"] == 1
    assert s["declined"] == 1
    assert s["pending"] == 1
    assert s["head_count"] == 3  # attendee + adult + child
    assert s["extra_adults"] == 1
    assert s["extra_children"] == 1
    # No explicit estimates: fallback chain gives each guest its tier's adult capacity
    # (plus_family 1+1=2, two solos 1+1) → 4.
    assert s["expected_head_count"] == 4

    # Generic per-question breakdown: the categorical "Main course" is tallied over
    # the one attending party (Fish ×1); the `number` Age question is not charted.
    bd = {b["prompt"]: b for b in s["question_breakdowns"]}
    assert "Age" not in bd
    course = bd["Main course"]
    assert course["scope"] == "invitee" and course["qtype"] == "choice"
    assert course["applicable"] == 1 and course["answered"] == 1
    assert course["counts"] == [{"label": "Fish", "count": 1}]


def test_summary_expected_head_count_sums_estimates(client, db_session, wedding):
    a = _add_guest(db_session, wedding, slug="e-1", name="Est A")
    b = _add_guest(db_session, wedding, slug="e-2", name="Est B")
    # No estimate set → falls back to the solo tier's capacity (1, just the invitee).
    _add_guest(db_session, wedding, slug="e-3", name="No Est")
    a.expected_party_size = 4
    b.expected_party_size = 2
    db_session.commit()
    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s["expected_head_count"] == 7


def test_summary_breakdown_person_multichoice(client, db_session, wedding):
    """A person-scope multi_choice (dietary) tallies each selected option across every
    attending person it applies to; `applicable`/`answered` reflect who could vs did."""
    diet = Question(
        wedding_id=wedding.id, prompt="Dietary", qtype=QuestionType.multi_choice,
        options=["Halal", "Vegetarian", "Nut allergy"],
        scope=QuestionScope.person, applies_to=QuestionApplies.everyone,
    )
    db_session.add(diet)
    db_session.commit()
    g = _add_guest(db_session, wedding, slug="d-1", name="Host", tier=InviteTier.plus_family)
    r = Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True)
    db_session.add(r)
    db_session.flush()
    plus = Companion(wedding_id=wedding.id, rsvp_id=r.id, kind=CompanionKind.adult, name="P1")
    kid = Companion(wedding_id=wedding.id, rsvp_id=r.id, kind=CompanionKind.child, name="K")
    r.companions[:] = [plus, kid]
    db_session.flush()
    r.answers[:] = [
        Answer(wedding_id=wedding.id, question_id=diet.id, companion_id=None,
               value={"choices": ["Halal"]}),
        Answer(wedding_id=wedding.id, question_id=diet.id, companion_id=plus.id,
               value={"choices": ["Halal", "Vegetarian"]}),
        # kid: left unanswered
    ]
    db_session.commit()
    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    bd = {b["prompt"]: b for b in s["question_breakdowns"]}["Dietary"]
    assert bd["scope"] == "person" and bd["applies_to"] == "everyone"
    assert bd["applicable"] == 3  # primary + adult + child are all asked
    assert bd["answered"] == 2  # the kid didn't answer
    assert bd["counts"] == [{"label": "Halal", "count": 2}, {"label": "Vegetarian", "count": 1}]


def test_summary_by_side(client, db_session, wedding):
    """by_side groups the list per guest side with each side's RSVP rollup; it's
    empty when nobody has a side, and pins 'Unassigned' last when some do."""
    # No sides yet → no split.
    assert client.get("/api/admin/summary", headers=dev_auth()).json()["by_side"] == []

    th = _add_guest(db_session, wedding, slug="s-th", name="Th One", tier=InviteTier.plus_one)
    th.side = "Alex"
    sh = _add_guest(db_session, wedding, slug="s-sh", name="Sh One")
    sh.side = "Sam"
    _add_guest(db_session, wedding, slug="s-none", name="No Side")  # → Unassigned
    r = Rsvp(wedding_id=wedding.id, guest_id=th.id, attending=True)
    db_session.add(r)
    db_session.flush()
    r.companions[:] = [Companion(wedding_id=wedding.id, rsvp_id=r.id, kind=CompanionKind.adult, name="P")]
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=sh.id, attending=False))
    db_session.commit()

    rows = client.get("/api/admin/summary", headers=dev_auth()).json()["by_side"]
    by = {row["label"]: row for row in rows}
    assert rows[-1]["label"] == "Unassigned"  # pinned last
    assert by["Alex"]["attending"] == 1 and by["Alex"]["head_count"] == 2  # primary + 1
    assert by["Sam"]["declined"] == 1 and by["Sam"]["head_count"] == 0
    assert by["Unassigned"]["pending"] == 1


def test_summary_capacity_and_invited_people(client, db_session, wedding):
    """The capacity-utilization lens: summary echoes the owner's `event_details.capacity`
    and `invited_people` tallies the expected size of invited-but-unreplied guests
    (overall and per side), excluding pending and declined."""
    # Default: no capacity configured.
    s0 = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s0["capacity"] == {"total": None, "by_side": {}}

    # One attending (head 2), one invited (expected 3), one pending, one declined — all Alex.
    att = _add_guest(db_session, wedding, slug="cap-att", name="Att One", tier=InviteTier.plus_one)
    att.side = "Alex"
    inv = _add_guest(db_session, wedding, slug="cap-inv", name="Inv One")
    inv.side = "Alex"
    inv.expected_party_size = 3
    pen = _add_guest(db_session, wedding, slug="cap-pen", name="Pen One")
    pen.side = "Sam"
    dec = _add_guest(db_session, wedding, slug="cap-dec", name="Dec One")
    dec.side = "Sam"
    dec.expected_party_size = 9  # declined → must NOT count toward invited_people
    r = Rsvp(wedding_id=wedding.id, guest_id=att.id, attending=True)
    db_session.add(r)
    db_session.flush()
    r.companions[:] = [Companion(wedding_id=wedding.id, rsvp_id=r.id, kind=CompanionKind.adult, name="P")]
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=dec.id, attending=False))
    db_session.commit()
    client.put(f"/api/admin/guests/{inv.id}/rsvp", headers=dev_auth(), json={"status": "invited"})

    # Set capacity via the content PATCH (deep-merged into event_details).
    client.patch(
        "/api/admin/content",
        headers=dev_auth(),
        json={"event_details": {"capacity": {"total": 120, "by_side": {"Alex": 60, "Sam": 50}}}},
    )

    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s["capacity"] == {"total": 120, "by_side": {"Alex": 60, "Sam": 50}}
    assert s["head_count"] == 2  # confirmed (attending primary + 1)
    assert s["invited_people"] == 3  # the invited guest's expected size; declined's 9 excluded

    by = {row["label"]: row for row in s["by_side"]}
    assert by["Alex"]["head_count"] == 2 and by["Alex"]["invited_people"] == 3
    assert by["Sam"]["head_count"] == 0 and by["Sam"]["invited_people"] == 0


def test_summary_capacity_ignores_garbage(client, db_session, wedding):
    """A malformed capacity blob (non-numeric / negative / wrong type) degrades to
    unset rather than 500-ing the summary."""
    client.patch(
        "/api/admin/content",
        headers=dev_auth(),
        json={"event_details": {"capacity": {"total": "lots", "by_side": {"Alex": -5, "": 10}}}},
    )
    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s["capacity"] == {"total": None, "by_side": {}}


def test_summary_pivot(client, db_session, wedding):
    """The pivot groups invitation status by any dimension, only offers dimensions that
    actually split the list, splits no-reply into invited vs pending, and nests one
    level via `then`."""
    # Two sides × two batches; Alex/batch-A is invited (link sent, no reply),
    # Alex/batch-B attending, Sam/batch-A pending (not contacted).
    g1 = _add_guest(db_session, wedding, slug="p1", name="A One")
    g1.side, g1.batch, g1.invite_sent = "Alex", "Batch A", True
    g2 = _add_guest(db_session, wedding, slug="p2", name="B Two")
    g2.side, g2.batch = "Alex", "Batch B"
    g3 = _add_guest(db_session, wedding, slug="p3", name="C Three")
    g3.side, g3.batch = "Sam", "Batch A"
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=g2.id, attending=True))
    db_session.commit()

    body = client.get("/api/admin/summary/pivot?by=side&then=batch", headers=dev_auth()).json()
    assert body["by"] == "side" and body["then"] == "batch"
    # status + side + batch are offered (3 statuses present; everyone is the same tier;
    # relationship/group unset → those are hidden).
    assert set(body["available_dims"]) == {"status", "side", "batch"}
    assert body["total"]["invitations"] == 3 and body["total"]["people"] >= 3
    groups = {row["label"]: row for row in body["groups"]}
    alex = groups["Alex"]
    assert alex["invitations"] == 2 and alex["invited"] == 1 and alex["attending"] == 1
    assert {c["label"] for c in alex["children"]} == {"Batch A", "Batch B"}
    assert groups["Sam"]["pending"] == 1

    # An invalid `by` falls back to "side"; a `then` equal to `by` drops to no stack.
    fb = client.get("/api/admin/summary/pivot?by=group&then=side", headers=dev_auth()).json()
    assert fb["by"] == "side" and fb["then"] is None

    # `then=status` is the default stack — children are the funnel-ordered statuses.
    st = client.get("/api/admin/summary/pivot?by=side&then=status", headers=dev_auth()).json()
    alex_st = next(r for r in st["groups"] if r["label"] == "Alex")
    assert [c["label"] for c in alex_st["children"]] == ["Attending", "Invited"]

    # The `side` filter scopes the data but leaves the dimension menu (full list) intact.
    alex_only = client.get("/api/admin/summary/pivot?by=side&side=Alex", headers=dev_auth()).json()
    assert "side" in alex_only["available_dims"]
    assert [r["label"] for r in alex_only["groups"]] == ["Alex"]
    assert alex_only["total"]["invitations"] == 2

    # The `status` filter (Confirmed tab) keeps only attending parties; people = heads.
    conf = client.get("/api/admin/summary/pivot?by=side&then=&status=attending", headers=dev_auth()).json()
    assert conf["total"]["invitations"] == 1 and conf["total"]["attending"] == 1
    assert conf["total"]["people"] == conf["total"]["head_count"]


def test_summary_timeline_and_week_windows(client, db_session, wedding):
    """replies_this_week/last_week count first replies in trailing 7/8-14 day windows;
    /summary/timeline buckets cumulative replies by week (counted by responded_at)."""
    from datetime import datetime, timedelta, timezone

    now = datetime.now(timezone.utc).replace(tzinfo=None)

    def add_reply(slug, days_ago):
        g = _add_guest(db_session, wedding, slug=slug, name=slug)
        db_session.add(
            Rsvp(
                wedding_id=wedding.id,
                guest_id=g.id,
                attending=True,
                responded_at=now - timedelta(days=days_ago),
            )
        )
        db_session.commit()

    add_reply("t-a", 1)   # this week
    add_reply("t-b", 3)   # this week
    add_reply("t-c", 10)  # last week (8-14d)
    add_reply("t-d", 22)  # older
    _add_guest(db_session, wedding, slug="t-pending", name="Pending")  # no RSVP

    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s["replies_this_week"] == 2
    assert s["replies_last_week"] == 1

    tl = client.get("/api/admin/summary/timeline", headers=dev_auth()).json()
    assert tl["total_invitations"] == 5  # 4 replied + 1 pending
    assert tl["total_replied"] == 4
    cums = [p["cumulative"] for p in tl["points"]]
    assert cums == sorted(cums)  # non-decreasing
    assert cums[-1] == 4  # ends at the total
    assert sum(p["new"] for p in tl["points"]) == 4


# --- RSVP audit trail ------------------------------------------------------
def _find_guest(client, gid):
    rows = client.get("/api/admin/guests", headers=dev_auth()).json()
    return next(row for row in rows if row["id"] == str(gid))


def test_audit_guest_reply_then_admin_override(client, db_session, wedding):
    """A guest's own reply stamps source=guest (no actor); a later owner override
    flips last_source/last_actor to admin while first_source stays 'guest'."""
    g = _add_guest(db_session, wedding, slug="aud-1", name="Aud One")
    assert client.post("/api/i/aud-1/rsvp", json={"attending": True}).status_code == 200

    row = _find_guest(client, g.id)
    assert row["first_source"] == "guest" and row["last_source"] == "guest"
    assert row["last_actor"] is None
    assert row["responded_at"] and row["updated_at"]

    res = client.put(
        f"/api/admin/guests/{g.id}/rsvp", json={"status": "declined"}, headers=dev_auth()
    )
    assert res.status_code == 200
    body = res.json()
    assert body["first_source"] == "guest"  # preserved — the guest replied first
    assert body["last_source"] == "admin" and body["last_actor"] == ADMIN_EMAIL


def test_audit_admin_create_and_bulk_source(client, db_session, wedding):
    """An RSVP first recorded by the owner is stamped source=admin; bulk too."""
    g = _add_guest(db_session, wedding, slug="aud-2", name="Aud Two")
    body = client.put(
        f"/api/admin/guests/{g.id}/rsvp", json={"status": "attending"}, headers=dev_auth()
    ).json()
    assert body["first_source"] == "admin" and body["last_source"] == "admin"
    assert body["last_actor"] == ADMIN_EMAIL

    g2 = _add_guest(db_session, wedding, slug="aud-3", name="Aud Three")
    client.post(
        "/api/admin/guests/bulk/rsvp",
        json={"ids": [str(g2.id)], "status": "attending"},
        headers=dev_auth(),
    )
    row = _find_guest(client, g2.id)
    assert row["last_source"] == "admin" and row["last_actor"] == ADMIN_EMAIL


def test_responses_detail(client, db_session, wedding):
    _seed_responses(db_session, wedding)
    rows = client.get("/api/admin/responses", headers=dev_auth()).json()
    attending = next(r for r in rows if r["attending"])
    assert attending["guest_name"] == "Attendee One"
    assert len(attending["companions"]) == 2
    # Party answer (the invitee-scope choice).
    assert attending["answers"][0]["prompt"] == "Main course"
    assert attending["answers"][0]["value"] == {"choice": "Fish"}
    # The child's per-person age answer rides on that companion.
    kid = next(c for c in attending["companions"] if c["name"] == "Kiddo")
    assert kid["answers"][0]["value"] == {"number": 4}


_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _read_xlsx(content: bytes):
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def test_export_and_template_xlsx(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="x-1", name="Xlsx Person", tier=InviteTier.plus_family)
    g.expected_party_size = 3
    rsvp = Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True)
    db_session.add(rsvp)
    db_session.flush()
    db_session.add(Companion(wedding_id=wedding.id, rsvp_id=rsvp.id, kind=CompanionKind.child, name="Kid"))
    db_session.commit()

    r = client.get("/api/admin/export.xlsx", headers=dev_auth())
    assert r.status_code == 200 and r.headers["content-type"].startswith(_XLSX_CT)
    rows = _read_xlsx(r.content)
    assert rows[0][:4] == ["Id", "Link", "Person", "Name"]  # no Invitee column
    person_col = rows[0].index("Person")
    # The Primary row carries the guest's UUID in the Id column.
    primary = next(row for row in rows[1:] if row[person_col] == "Primary")
    assert primary[0] == str(g.id)
    # The Expected column round-trips on the Primary row.
    expected_col = rows[0].index("Expected")
    assert str(primary[expected_col]) == "3"
    # The Actual column shows the computed party size (attending primary + 1 child).
    actual_col = rows[0].index("Actual")
    assert str(primary[actual_col]) == "2"
    # Primary + one child row for the family guest.
    bodies = [row[person_col] for row in rows[1:]]
    assert "Primary" in bodies and "Child" in bodies

    t = client.get("/api/admin/template.xlsx", headers=dev_auth())
    assert t.status_code == 200 and t.headers["content-type"].startswith(_XLSX_CT)
    trows = _read_xlsx(t.content)
    assert trows[0][:4] == ["Id", "Link", "Person", "Name"]


def test_export_xlsx_has_dropdowns(client, db_session, wedding):
    """The workbook carries data-validation dropdowns (a hidden Lists sheet + at
    least one list validation on the Guests sheet)."""
    import io

    import openpyxl

    _add_guest(db_session, wedding, slug="dv-1", name="DV Person")
    db_session.commit()
    r = client.get("/api/admin/export.xlsx", headers=dev_auth())
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Lists" in wb.sheetnames
    assert len(wb["Guests"].data_validations.dataValidation) > 0


def _csv_upload(text: str):
    return {"file": ("guests.csv", text.encode("utf-8"), "text/csv")}


def test_import_dry_run_then_commit(client, db_session, wedding):
    csv_text = (
        "Id,Link,Invitee,Person,Name,Greeting,Email,Phone,Tier,Attending\n"
        # An Id that doesn't resolve → error row (Link is display-only, not matched).
        "00000000-0000-0000-0000-000000000000,,Ghost,Primary,Ghost,Ghost,,,,\n"
        ",,Fresh Family,Primary,Fresh Family,Fresh Family,Me@Example.COM,91234567,plus_family,yes\n"
        ",,Fresh Family,Child,Leo,,,,,\n"
    )
    # Dry run: previews, persists nothing.
    dry = client.post("/api/admin/import?commit=0", headers=dev_auth(), files=_csv_upload(csv_text))
    assert dry.status_code == 200
    body = dry.json()
    assert body["committed"] is False
    assert body["created"] == 1 and body["errors"] == 1
    assert db_session.query(Guest).count() == 0

    # Commit: creates the new guest with normalized contacts + the child RSVP.
    done = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert done.status_code == 200 and done.json()["committed"] is True
    g = db_session.query(Guest).filter(Guest.name == "Fresh Family").one()
    assert g.email == "Me@example.com" and g.phone == "+6591234567"
    assert g.invite_tier == InviteTier.plus_family
    assert g.rsvp.attending is True
    kids = [c for c in g.rsvp.companions if c.kind == CompanionKind.child]
    assert len(kids) == 1 and kids[0].name == "Leo"
    # Audit: an import is stamped as source=import with the owner as actor.
    assert g.rsvp.first_source == "import" and g.rsvp.last_source == "import"
    assert g.rsvp.last_actor == ADMIN_EMAIL


def test_import_expected_party_size(client, db_session, wedding):
    # A valid Expected value imports; an out-of-range one is reported as an error.
    csv_text = (
        "Id,Invitee,Person,Name,Greeting,Tier,Expected\n"
        ",Planned Party,Primary,Planned Party,Planned Party,plus_family,5\n"
        ",Too Big,Primary,Too Big,Too Big,solo,999\n"
    )
    dry = client.post("/api/admin/import?commit=0", headers=dev_auth(), files=_csv_upload(csv_text))
    assert dry.status_code == 200
    assert dry.json()["errors"] == 1  # the 999 row

    client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    g = db_session.query(Guest).filter(Guest.name == "Planned Party").one()
    assert g.expected_party_size == 5
    # The invalid row was skipped, so no "Too Big" guest exists.
    assert db_session.query(Guest).filter(Guest.name == "Too Big").count() == 0


def test_create_guest_stores_greeting_name(client, wedding):
    r = client.post(
        "/api/admin/guests",
        headers=dev_auth(),
        json={"name": "John Smith", "invite_tier": "plus_one", "greeting_name": "John & Jane"},
    )
    assert r.status_code == 201
    assert r.json()["greeting_name"] == "John & Jane"


def test_create_guest_requires_greeting(client, wedding):
    # Greeting is mandatory: a missing or empty greeting is a 422; the name is optional.
    missing = client.post("/api/admin/guests", headers=dev_auth(), json={"name": "No Greet"})
    assert missing.status_code == 422
    empty = client.post(
        "/api/admin/guests", headers=dev_auth(), json={"greeting_name": ""}
    )
    assert empty.status_code == 422
    # Name-less invite with only a greeting is valid; slug is a non-descript token.
    ok = client.post("/api/admin/guests", headers=dev_auth(), json={"greeting_name": "John & Jane"})
    assert ok.status_code == 201
    body = ok.json()
    assert body["name"] == "" and body["greeting_name"] == "John & Jane"
    assert all(c.isalnum() or c in "-_" for c in body["slug"]) and "john" not in body["slug"]


def test_party_members_persist_and_clamp_to_tier(client, db_session, wedding):
    # A plus_one invite keeps one adult; a stray child is dropped to the tier cap.
    r = client.post(
        "/api/admin/guests",
        headers=dev_auth(),
        json={
            "greeting_name": "Sam & Alex",
            "invite_tier": "plus_one",
            "party_members": [
                {"kind": "adult", "name": "Alex"},
                {"kind": "child", "name": "Nope"},
            ],
        },
    )
    assert r.status_code == 201
    gid = r.json()["id"]
    assert r.json()["party_members"] == [{"kind": "adult", "name": "Alex"}]

    # Updating the prefill party replaces it (still clamped).
    upd = client.patch(
        f"/api/admin/guests/{gid}",
        headers=dev_auth(),
        json={"party_members": [{"kind": "adult", "name": "Alexandra"}]},
    )
    assert upd.json()["party_members"] == [{"kind": "adult", "name": "Alexandra"}]


def test_set_attending_materializes_prefill_party_as_companions(client, db_session, wedding):
    # Marking a guest attending turns their admin-prefilled +1/kids (party_members)
    # into real companions, so per-person answers (age, dietary) become editable.
    g = _add_guest(db_session, wedding, slug="mat-1", name="Lead", tier=InviteTier.plus_family)
    g.party_members = [{"kind": "adult", "name": "Robin"}, {"kind": "child", "name": "Junior"}]
    db_session.commit()

    r = client.put(f"/api/admin/guests/{g.id}/rsvp", headers=dev_auth(), json={"status": "attending"})
    assert r.status_code == 200
    comps = {c["name"]: c["kind"] for c in r.json()["companions"]}
    assert comps == {"Robin": "adult", "Junior": "child"}


def test_set_attending_keeps_existing_party(client, db_session, wedding):
    # If an RSVP already has a party (e.g. the guest responded), re-marking attending
    # does NOT clobber it with the prefill.
    g = _add_guest(db_session, wedding, slug="mat-2", name="Lead", tier=InviteTier.plus_one)
    g.party_members = [{"kind": "adult", "name": "Prefill"}]
    rsvp = Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True)
    rsvp.companions.append(Companion(wedding_id=wedding.id, kind=CompanionKind.adult, name="Real"))
    db_session.add(rsvp)
    db_session.commit()

    r = client.put(f"/api/admin/guests/{g.id}/rsvp", headers=dev_auth(), json={"status": "attending"})
    assert r.status_code == 200
    assert [c["name"] for c in r.json()["companions"]] == ["Real"]


def test_import_seeds_party_members_without_rsvp(client, db_session, wedding):
    # Adult/Child Name rows seed the prefill party even when Attending is blank, so the
    # guest's RSVP opens with names ready (no RSVP is created).
    csv_text = (
        "Id,Invitee,Person,Name,Greeting,Tier\n"
        ",Lead,Primary,Lead,Lead & Co,plus_family\n"
        ",Lead,Adult,Robin,,\n"
        ",Lead,Child,Junior,,\n"
    )
    r = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200 and r.json()["created"] == 1
    g = db_session.query(Guest).filter(Guest.greeting_name == "Lead & Co").one()
    assert g.rsvp is None  # Attending blank → no RSVP
    assert g.party_members == [
        {"kind": "adult", "name": "Robin"},
        {"kind": "child", "name": "Junior"},
    ]


def test_import_and_export_greeting_name(client, db_session, wedding):
    # Greeting is invitee-level: it imports off the Primary row only; a value on a
    # companion row is not treated as a field. It then round-trips back on export.
    csv_text = (
        "Id,Invitee,Person,Name,Greeting,Tier\n"
        ",John,Primary,John,John & Jane,plus_one\n"
        ",John,Adult,Jane,ignored,\n"
    )
    r = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200 and r.json()["created"] == 1
    g = db_session.query(Guest).filter(Guest.name == "John").one()
    assert g.greeting_name == "John & Jane"

    # Export: Greeting populated on the Primary row, blank on companion rows.
    grid = _read_xlsx(client.get("/api/admin/export.xlsx", headers=dev_auth()).content)
    header = [str(h) for h in grid[0]]
    rows = [dict(zip(header, [("" if v is None else str(v)) for v in row])) for row in grid[1:]]
    assert "Greeting" in header
    primary = next(row for row in rows if row["Person"] == "Primary")
    assert primary["Greeting"] == "John & Jane"


def test_import_updates_by_id(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="upd-1", name="Update Me", tier=InviteTier.solo)
    db_session.commit()
    gid = str(g.id)
    csv_text = (
        "Id,Invitee,Person,Name,Greeting,Email,Tier\n"
        f"{gid},Update Me,Primary,Update Me,Update Me,new@example.com,plus_one\n"
    )
    r = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200
    body = r.json()
    assert body["updated"] == 1 and body["created"] == 0
    assert db_session.query(Guest).count() == 1  # updated in place, not duplicated
    db_session.refresh(g)
    assert g.email == "new@example.com" and g.invite_tier == InviteTier.plus_one


def test_import_writes_admin_answers(client, db_session, wedding):
    """Admin-defined question columns round-trip in: a per-person dietary + a
    children-only age land as Answers on the right person."""
    diet_q = Question(
        wedding_id=wedding.id, prompt="Any dietary needs?", qtype=QuestionType.multi_choice,
        options=["Halal", "Vegetarian", "Vegan"], scope=QuestionScope.person,
        applies_to=QuestionApplies.everyone,
    )
    age_q = Question(
        wedding_id=wedding.id, prompt="Age", qtype=QuestionType.number,
        scope=QuestionScope.person, applies_to=QuestionApplies.children,
    )
    db_session.add_all([diet_q, age_q])
    db_session.commit()
    csv_text = (
        "Id,Invitee,Person,Name,Greeting,Tier,Attending,Any dietary needs?,Age\n"
        ",Imp Fam,Primary,Imp Fam,Imp Fam,plus_family,yes,Halal,\n"
        ',Imp Fam,Child,Lee,,,,"Vegan, Vegetarian",7\n'
    )
    r = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200 and r.json()["created"] == 1
    g = db_session.query(Guest).filter(Guest.name == "Imp Fam").one()
    # Primary's own dietary (party answer, companion_id None).
    party = [a for a in g.rsvp.answers if a.companion_id is None]
    assert any(a.question_id == diet_q.id and a.value == {"choices": ["Halal"]} for a in party)
    # The child's dietary (multi) + age ride on the child companion.
    child = next(c for c in g.rsvp.companions if c.kind == CompanionKind.child)
    cvals = {a.question_id: a.value for a in child.answers}
    assert cvals[diet_q.id] == {"choices": ["Vegan", "Vegetarian"]}
    assert cvals[age_q.id] == {"number": 7}


def test_import_counts_invitees_and_people(client, db_session, wedding):
    """The dry-run reports invitee counts AND person (incl. companion) counts."""
    csv_text = (
        "Id,Person,Name,Greeting,Tier\n"
        ",Primary,Adam,Adam,plus_family\n"
        ",Adult,Mum,,\n"
        ",Adult,Wife,,\n"
        ",Child,Kid,,\n"
        ",Primary,Sam,Sam,solo\n"
    )
    r = client.post("/api/admin/import", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200
    body = r.json()
    assert body["committed"] is False
    assert body["created"] == 2  # two invitees (Adam's family + Sam)
    assert body["people_created"] == 5  # Adam + mum + wife + kid (4) + Sam (1)
    assert body["updated"] == 0 and body["people_updated"] == 0


def test_import_rejects_bad_answer_value(client, db_session, wedding):
    """A non-numeric Age cell is reported as a row error, not committed."""
    db_session.add(
        Question(
            wedding_id=wedding.id, prompt="Age", qtype=QuestionType.number,
            scope=QuestionScope.person, applies_to=QuestionApplies.children,
        )
    )
    db_session.commit()
    csv_text = (
        "Id,Invitee,Person,Name,Tier,Attending,Age\n"
        ",Bad Fam,Primary,Bad Fam,plus_family,yes,\n"
        ",Bad Fam,Child,Kid,,,not-a-number\n"
    )
    r = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200
    body = r.json()
    assert body["errors"] == 1 and body["created"] == 0
    assert "number" in body["rows"][0]["detail"]
    assert db_session.query(Guest).count() == 0


def test_export_per_person_and_invitee_copydown(client, db_session, wedding):
    """Companion rows carry their OWN person-scope answers (dietary/age), while
    invitee-scope answers + the invitee contact fields are copied DOWN onto them."""
    g, adult, child, diet_q, _age_q = _party_with_companions(db_session, wedding)
    g.email = "host@example.com"
    know_q = Question(
        wedding_id=wedding.id, prompt="How?", qtype=QuestionType.text,
        scope=QuestionScope.invitee, applies_to=QuestionApplies.everyone, sort_order=5,
    )
    db_session.add(know_q)
    db_session.flush()
    db_session.add_all([
        # party (invitee-scope) answer
        Answer(wedding_id=wedding.id, rsvp_id=g.rsvp.id, question_id=know_q.id, value={"text": "Friends"}),
        # per-person dietary: primary (companion_id None), adult, child
        Answer(wedding_id=wedding.id, rsvp_id=g.rsvp.id, question_id=diet_q.id, value={"choices": ["Halal"]}),
        Answer(wedding_id=wedding.id, rsvp_id=g.rsvp.id, question_id=diet_q.id, companion_id=adult.id, value={"choices": ["Vegetarian"]}),
        Answer(wedding_id=wedding.id, rsvp_id=g.rsvp.id, question_id=diet_q.id, companion_id=child.id, value={"choices": ["Vegan"]}),
    ])
    db_session.commit()

    r = client.get("/api/admin/export.xlsx", headers=dev_auth())
    grid = _read_xlsx(r.content)
    header = [str(h) for h in grid[0]]
    rows = [dict(zip(header, [("" if v is None else str(v)) for v in row])) for row in grid[1:]]
    primary = next(row for row in rows if row["Person"] == "Primary")
    adult_r = next(row for row in rows if row["Person"] == "Guest")
    child_r = next(row for row in rows if row["Person"] == "Child")

    # Per-person dietary shows each person's OWN value.
    assert primary["Any dietary needs?"] == "Halal"
    assert adult_r["Any dietary needs?"] == "Vegetarian"
    assert child_r["Any dietary needs?"] == "Vegan"
    # Age (person-scope, children-only) only on the child row.
    assert child_r["Age"] == "6" and adult_r["Age"] == ""
    # Invitee-scope answer copied down to every companion row.
    assert primary["How?"] == adult_r["How?"] == child_r["How?"] == "Friends"
    # Invitee contact copied down; Link stays Primary-only.
    assert adult_r["Email"] == child_r["Email"] == "host@example.com"
    assert primary["Link"] and adult_r["Link"] == "" and child_r["Link"] == ""
    # Id + Greeting are copied onto EVERY party row (the Id ties them to the invite).
    assert primary["Id"] == adult_r["Id"] == child_r["Id"] == str(g.id)
    assert primary["Greeting"] == adult_r["Greeting"] == child_r["Greeting"]


def test_set_guest_rsvp_attending_with_answers(client, db_session, wedding):
    know_q = Question(
        wedding_id=wedding.id, prompt="How?", qtype=QuestionType.text, scope=QuestionScope.invitee,
    )
    diet_q = Question(
        wedding_id=wedding.id, prompt="Diet", qtype=QuestionType.text,
        scope=QuestionScope.person, applies_to=QuestionApplies.everyone,
    )
    db_session.add_all([know_q, diet_q])
    g = _add_guest(db_session, wedding, slug="rs-1", name="Resi")
    db_session.commit()

    r = client.put(
        f"/api/admin/guests/{g.id}/rsvp",
        headers=dev_auth(),
        json={
            "status": "attending",
            "answers": [
                {"question_id": str(know_q.id), "value": {"text": "Friends"}},
                {"question_id": str(diet_q.id), "value": {"text": "Halal"}},
            ],
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["rsvp_status"] == "attending"
    vals = {a["prompt"]: a["value"] for a in body["answers"]}
    assert vals["How?"] == {"text": "Friends"} and vals["Diet"] == {"text": "Halal"}


def test_set_guest_rsvp_pending_clears_rsvp(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="rs-2", name="Penny")
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True))
    db_session.commit()

    r = client.put(f"/api/admin/guests/{g.id}/rsvp", headers=dev_auth(), json={"status": "pending"})
    assert r.status_code == 200 and r.json()["rsvp_status"] == "pending"
    db_session.expire_all()
    assert db_session.query(Rsvp).filter(Rsvp.guest_id == g.id).count() == 0


def test_set_guest_rsvp_invited_marks_sent_without_rsvp(client, db_session, wedding):
    """`invited` is the Pending→Invited step: no RSVP row, but invite_sent flips on."""
    g = _add_guest(db_session, wedding, slug="inv-1", name="Sent")
    db_session.commit()
    # Pending by default.
    body = client.get("/api/admin/guests", headers=dev_auth()).json()
    me = next(x for x in body if x["slug"] == "inv-1")
    assert me["rsvp_status"] == "pending" and me["invite_sent"] is False

    r = client.put(f"/api/admin/guests/{g.id}/rsvp", headers=dev_auth(), json={"status": "invited"})
    assert r.status_code == 200
    assert r.json()["rsvp_status"] == "invited" and r.json()["invite_sent"] is True
    db_session.expire_all()
    # No RSVP row was created — "invited" is still "no reply yet".
    assert db_session.query(Rsvp).filter(Rsvp.guest_id == g.id).count() == 0

    # Back to pending clears the sent flag.
    r2 = client.put(f"/api/admin/guests/{g.id}/rsvp", headers=dev_auth(), json={"status": "pending"})
    assert r2.json()["rsvp_status"] == "pending" and r2.json()["invite_sent"] is False


def test_bulk_set_invited(client, db_session, wedding):
    a = _add_guest(db_session, wedding, slug="bi-1", name="Ann")
    b = _add_guest(db_session, wedding, slug="bi-2", name="Ben")
    db_session.commit()
    r = client.post(
        "/api/admin/guests/bulk/rsvp",
        headers=dev_auth(),
        json={"ids": [str(a.id), str(b.id)], "status": "invited"},
    )
    assert r.status_code == 200 and r.json()["count"] == 2
    rows = {x["slug"]: x for x in client.get("/api/admin/guests", headers=dev_auth()).json()}
    assert rows["bi-1"]["rsvp_status"] == "invited" and rows["bi-2"]["rsvp_status"] == "invited"


def test_summary_counts_invited_separately(client, db_session, wedding):
    _add_guest(db_session, wedding, slug="su-att", name="A")
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=db_session.query(Guest).filter_by(slug="su-att").one().id, attending=True))
    inv = _add_guest(db_session, wedding, slug="su-inv", name="I")
    _add_guest(db_session, wedding, slug="su-pen", name="P")  # pending
    db_session.commit()
    client.put(f"/api/admin/guests/{inv.id}/rsvp", headers=dev_auth(), json={"status": "invited"})

    s = client.get("/api/admin/summary", headers=dev_auth()).json()
    assert s["total_guests"] == 3
    assert s["attending"] == 1 and s["invited"] == 1 and s["pending"] == 1
    assert s["declined"] == 0


def test_set_guest_rsvp_sets_full_party_with_companion_answers(client, db_session, wedding):
    # The owner can set the whole attending party in one call: the primary's party
    # answers + each companion (name + its own person answers, e.g. a child's age).
    diet_q = Question(
        wedding_id=wedding.id, prompt="Diet", qtype=QuestionType.text,
        scope=QuestionScope.person, applies_to=QuestionApplies.everyone,
    )
    age_q = Question(
        wedding_id=wedding.id, prompt="Age", qtype=QuestionType.number,
        scope=QuestionScope.person, applies_to=QuestionApplies.children,
    )
    db_session.add_all([diet_q, age_q])
    g = _add_guest(db_session, wedding, slug="fp-1", name="Lead", tier=InviteTier.plus_family)
    db_session.commit()

    r = client.put(
        f"/api/admin/guests/{g.id}/rsvp",
        headers=dev_auth(),
        json={
            "status": "attending",
            "answers": [{"question_id": str(diet_q.id), "value": {"text": "Halal"}}],
            "companions": [
                {"kind": "adult", "name": "Robin", "answers": [
                    {"question_id": str(diet_q.id), "value": {"text": "Vegan"}}]},
                {"kind": "child", "name": "Junior", "answers": [
                    {"question_id": str(age_q.id), "value": {"number": 7}},
                    {"question_id": str(diet_q.id), "value": {"text": "No nuts"}}]},
            ],
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["rsvp_status"] == "attending" and body["party_size"] == 3
    # Primary's party answer.
    assert {a["prompt"]: a["value"] for a in body["answers"]}["Diet"] == {"text": "Halal"}
    comps = {c["name"]: {a["prompt"]: a["value"] for a in c["answers"]} for c in body["companions"]}
    assert comps["Robin"]["Diet"] == {"text": "Vegan"}
    assert comps["Junior"]["Age"] == {"number": 7} and comps["Junior"]["Diet"] == {"text": "No nuts"}


def test_set_guest_rsvp_companions_rejected_over_tier(client, db_session, wedding):
    # A solo invite can't be given companions via the RSVP override (tier cap).
    g = _add_guest(db_session, wedding, slug="fp-2", name="Solo", tier=InviteTier.solo)
    db_session.commit()
    r = client.put(
        f"/api/admin/guests/{g.id}/rsvp",
        headers=dev_auth(),
        json={"status": "attending", "companions": [{"kind": "adult", "name": "Nope", "answers": []}]},
    )
    assert r.status_code == 422


def test_set_guest_rsvp_rejects_child_only_question(client, db_session, wedding):
    """A children-only question is not a party question — answering it for the
    (adult) primary via the admin RSVP edit is rejected."""
    age_q = Question(
        wedding_id=wedding.id, prompt="Age", qtype=QuestionType.number,
        scope=QuestionScope.person, applies_to=QuestionApplies.children,
    )
    db_session.add(age_q)
    g = _add_guest(db_session, wedding, slug="rs-3", name="Adam")
    db_session.commit()
    r = client.put(
        f"/api/admin/guests/{g.id}/rsvp",
        headers=dev_auth(),
        json={"status": "attending", "answers": [{"question_id": str(age_q.id), "value": {"number": 9}}]},
    )
    assert r.status_code == 422


# --- Bulk guest actions ----------------------------------------------------
def test_bulk_set_rsvp_status(client, db_session, wedding):
    a = _add_guest(db_session, wedding, slug="b-1", name="Aned")
    b = _add_guest(db_session, wedding, slug="b-2", name="Bned")
    # b already attending with a companion — declining must drop the party.
    rsvp = Rsvp(wedding_id=wedding.id, guest_id=b.id, attending=True)
    db_session.add(rsvp)
    db_session.flush()
    db_session.add(Companion(wedding_id=wedding.id, rsvp_id=rsvp.id, kind=CompanionKind.adult, name="Plus"))
    db_session.commit()

    r = client.post(
        "/api/admin/guests/bulk/rsvp",
        headers=dev_auth(),
        json={"ids": [str(a.id), str(b.id)], "status": "declined"},
    )
    assert r.status_code == 200 and r.json()["count"] == 2
    rows = {row["name"]: row for row in client.get("/api/admin/guests", headers=dev_auth()).json()}
    assert rows["Aned"]["rsvp_status"] == "declined"
    assert rows["Bned"]["rsvp_status"] == "declined" and rows["Bned"]["party_size"] == 0


def test_bulk_set_rsvp_pending_clears(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="b-3", name="Pendable")
    db_session.add(Rsvp(wedding_id=wedding.id, guest_id=g.id, attending=True))
    db_session.commit()
    r = client.post(
        "/api/admin/guests/bulk/rsvp",
        headers=dev_auth(),
        json={"ids": [str(g.id)], "status": "pending"},
    )
    assert r.status_code == 200 and r.json()["count"] == 1
    db_session.expire_all()
    assert db_session.query(Rsvp).filter(Rsvp.guest_id == g.id).count() == 0


def test_bulk_delete_guests(client, db_session, wedding):
    a = _add_guest(db_session, wedding, slug="bd-1", name="Del One")
    b = _add_guest(db_session, wedding, slug="bd-2", name="Del Two")
    keep = _add_guest(db_session, wedding, slug="bd-3", name="Keep Me")
    r = client.post(
        "/api/admin/guests/bulk/delete",
        headers=dev_auth(),
        json={"ids": [str(a.id), str(b.id)]},
    )
    assert r.status_code == 200 and r.json()["count"] == 2
    names = [row["name"] for row in client.get("/api/admin/guests", headers=dev_auth()).json()]
    assert names == ["Keep Me"]
    assert db_session.query(Guest).filter(Guest.id == keep.id).count() == 1


def test_bulk_actions_are_tenant_scoped(make_client, db_session):
    """Foreign ids are silently excluded (count = only the owned ones), so a bulk
    call can't reach another wedding's guests."""
    a = Wedding(slug="wed-a", couple_names="A", status="active", owner_id="dev",
                event_details={}, content={})
    b = Wedding(slug="wed-b", couple_names="B", status="active", owner_id="other",
                event_details={}, content={})
    db_session.add_all([a, b])
    db_session.commit()
    ga = _add_guest(db_session, a, slug="ga", name="Guest A")
    gb = _add_guest(db_session, b, slug="gb", name="Guest B")
    client = make_client()  # dev owner → wedding A

    # Bulk delete a mix: only A's guest is removed; B's is untouched.
    r = client.post(
        "/api/admin/guests/bulk/delete",
        headers=dev_auth(),
        json={"ids": [str(ga.id), str(gb.id)]},
    )
    assert r.status_code == 200 and r.json()["count"] == 1
    assert db_session.query(Guest).filter(Guest.id == gb.id).count() == 1
    assert db_session.query(Guest).filter(Guest.id == ga.id).count() == 0


def test_bulk_rsvp_rejects_bad_status(client, db_session, wedding):
    g = _add_guest(db_session, wedding, slug="bb-1", name="Bad Status")
    r = client.post(
        "/api/admin/guests/bulk/rsvp",
        headers=dev_auth(),
        json={"ids": [str(g.id)], "status": "maybe"},
    )
    assert r.status_code == 422


def test_import_reports_over_tier(client, db_session, wedding):
    csv_text = (
        "Person,Name,Greeting,Tier,Attending\n"
        "Primary,Solo Sam,Solo Sam,solo,yes\n"
        "Adult,Plus One,,,\n"  # solo can't bring an adult
    )
    r = client.post("/api/admin/import?commit=1", headers=dev_auth(), files=_csv_upload(csv_text))
    assert r.status_code == 200
    body = r.json()
    assert body["errors"] == 1 and body["created"] == 0
    assert "exceed" in body["rows"][0]["detail"]
    assert db_session.query(Guest).count() == 0  # nothing committed for the bad row
