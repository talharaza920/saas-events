"""Wedding-scoped admin API (dashboard). Every endpoint lives under
`/api/w/{wedding_slug}/admin/*`; the tenant is resolved FROM THE PATH and access
is granted by an active membership row (or platform admin) — see app/authz.py.

  GET    …/me                → whoami + role + lifecycle + entitlements
  CRUD   …/guests, …/questions, …/story-arcs, …/wishes, …/content
  GET    …/responses, …/summary*   → RSVP rollups
  GET    …/export.xlsx, …/template.xlsx; POST …/import
  PUT    …/guests/{id}/rsvp  → owner override of a guest's RSVP
  POST   …/submit-approval   → draft → pending_approval (auto-rules may activate)
  POST   …/publish           → toggle publication (active weddings only)
  PATCH  …/settings          → per-wedding admin settings (owner)
  DELETE …/                  → soft-delete (archive) the wedding (owner)
  CRUD   …/members           → co-admin management (owner; Phase 3)

Reads stay available on a suspended wedding; mutations 403 (read-only mode).
Unlike the guest API, the admin DOES see/set `invite_tier` — the owner authors
the invites. Every query still filters by `wedding_id` (belt-and-braces under
the path-based scoping).
"""
from __future__ import annotations

import csv
import hashlib
import io
import secrets
from datetime import datetime, timedelta, timezone
from uuid import UUID

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app import export_import
from app.answers import is_party_question, person_question_applies
from app.approval import evaluate_auto_approval
from app.audit import SOURCE_ADMIN, SOURCE_IMPORT, stamp_rsvp
from app.audit_log import record
from app.authz import WeddingCtx, require_wedding
from app.config import Settings, get_settings
from app.db import get_db
from app.emailer import send_email
from app.entitlements import check_limit, check_storage, effective_entitlements, require_feature
from app.guest_import import make_guest_slug
from app.models import (
    Answer,
    Companion,
    CompanionKind,
    Guest,
    InviteTier,
    MemberRole,
    MemberStatus,
    Profile,
    Question,
    QuestionApplies,
    QuestionScope,
    QuestionType,
    QuestionVisibility,
    Rsvp,
    StoryArc,
    Wedding,
    WeddingMember,
    WeddingStatus,
    Wish,
)
from app.tenancy import capabilities_for, clamp_party_members
from app.schemas import (
    AdminMe,
    LifecycleResult,
    MemberAdmin,
    MemberInviteCreate,
    MemberInvited,
    MemberRoleUpdate,
    PublishUpdate,
    RuleTraceEntry,
    WeddingSettingsUpdate,
    AdminSummary,
    AnswerAdmin,
    BulkGuestIds,
    BulkResult,
    BulkRsvpUpdate,
    CapacityConfig,
    CompanionAdmin,
    CompanionUpdate,
    ContentAdmin,
    ContentUpdate,
    GroupBreakdown,
    GuestAdmin,
    GuestCreate,
    GuestRsvpUpdate,
    GuestUpdate,
    ImportResult,
    ImportRowResult,
    OptionCount,
    PivotSummary,
    QuestionAdmin,
    QuestionBreakdown,
    QuestionCreate,
    QuestionUpdate,
    ResponseAdmin,
    StoryArcAdmin,
    TimelinePoint,
    TimelineSummary,
    StoryArcCreate,
    StoryArcUpdate,
    UploadResult,
    WishAdmin,
    WishModerate,
)
from app.storage import UploadError, prepare_image, store_image
from app.validation import ContactError, normalize_email, normalize_phone, wedding_phone_region

router = APIRouter(prefix="/api/w/{wedding_slug}/admin", tags=["admin"])


# --- Tenant resolution (the authz seam, SAAS_PLAN 1.3) -----------------------
# Three grades of access, all resolving the wedding from the path:
#   member_ctx / member_wedding — any active member; works even when suspended
#   editor_ctx / editable_wedding — any active member; 403 while suspended
#   owner_ctx — owner-only mutations (members, lifecycle, delete)
member_ctx = require_wedding("admin")
editor_ctx = require_wedding("admin", edit=True)
owner_ctx = require_wedding("owner", edit=True)


def member_wedding(ctx: WeddingCtx = Depends(member_ctx)) -> Wedding:
    return ctx.wedding


def editable_wedding(ctx: WeddingCtx = Depends(editor_ctx)) -> Wedding:
    return ctx.wedding


def _validated_arc_ids(db: Session, wedding: Wedding, arc_ids) -> list[str]:
    """Normalize a guest's story-arc override to id strings, rejecting any arc that
    doesn't belong to this wedding (blocks cross-tenant / bogus targeting). Empty
    → []; the caller stores that as NULL so the guest falls back to all visible arcs."""
    if not arc_ids:
        return []
    owned = {
        str(x)
        for x in db.execute(
            select(StoryArc.id).where(StoryArc.wedding_id == wedding.id)
        ).scalars()
    }
    out = [str(a) for a in arc_ids]
    if any(a not in owned for a in out):
        raise HTTPException(status_code=422, detail="Unknown story arc")
    return out


def _unique_slug(db: Session, name: str) -> str:
    """A globally-unique, unguessable guest slug (retry on the rare collision)."""
    for _ in range(5):
        slug = make_guest_slug(name)
        if db.execute(select(Guest).where(Guest.slug == slug)).scalar_one_or_none() is None:
            return slug
    raise HTTPException(status_code=500, detail="Could not generate a unique link")


# --- Serializers -----------------------------------------------------------
def _answer_admins(answers, qmeta: dict) -> list[AnswerAdmin]:
    """Serialize Answer rows to AnswerAdmin, labelling each with its question's
    prompt + type. `qmeta` maps question_id → (prompt, qtype)."""
    out: list[AnswerAdmin] = []
    for a in answers:
        prompt, qtype = qmeta.get(a.question_id, ("(deleted question)", "text"))
        out.append(AnswerAdmin(question_id=a.question_id, prompt=prompt, qtype=qtype, value=a.value))
    return out


def _guest_admin(guest: Guest, qmeta: dict | None = None, content: dict | None = None) -> GuestAdmin:
    """Serialize a guest + its RSVP rollup. `qmeta` maps question_id → (prompt,
    qtype) so embedded answers carry their label/type; pass the wedding's question
    map when listing (a freshly-created guest has no answers, so {} is fine there).
    `content` is the wedding's content JSON, threaded so plus_family prefill clamping
    uses the wedding's configured (not default) companion caps.

    Party answers (invitee-scope + the primary's own person answers) have no
    companion_id; each companion's answers carry its id and ride on that companion.
    """
    qmeta = qmeta or {}
    rsvp = guest.rsvp
    notes = None
    companions: list[CompanionAdmin] = []
    answers: list[AnswerAdmin] = []
    updated_at = first_source = last_source = last_actor = None
    if rsvp is None:
        # No reply yet: "invited" once the owner has sent the invite, else "pending".
        status_str = "invited" if guest.invite_sent else "pending"
        party = 0
        responded_at = None
    else:
        notes = rsvp.notes
        responded_at = rsvp.responded_at
        updated_at = rsvp.updated_at
        first_source = rsvp.first_source
        last_source = rsvp.last_source
        last_actor = rsvp.last_actor
        if rsvp.attending:
            status_str = "attending"
            party = 1 + len(rsvp.companions)
        else:
            status_str = "declined"
            party = 0
        companions = [
            CompanionAdmin(
                id=c.id, kind=c.kind.value, name=c.name, answers=_answer_admins(c.answers, qmeta)
            )
            for c in rsvp.companions
        ]
        answers = _answer_admins([a for a in rsvp.answers if a.companion_id is None], qmeta)
    return GuestAdmin(
        id=guest.id,
        slug=guest.slug,
        invite_path=f"/i/{guest.slug}",
        name=guest.name,
        greeting_name=guest.greeting_name,
        party_members=clamp_party_members(guest.party_members, guest.invite_tier, content),
        email=guest.email,
        phone=guest.phone,
        side=guest.side,
        relationship=guest.relationship_label,
        group_name=guest.group_name,
        batch=guest.batch,
        invite_tier=guest.invite_tier.value,
        invited=guest.invited,
        invite_sent=guest.invite_sent,
        expected_party_size=guest.expected_party_size,
        story_arc_ids=guest.story_arc_ids or [],
        rsvp_status=status_str,
        party_size=party,
        notes=notes,
        responded_at=responded_at,
        updated_at=updated_at,
        first_source=first_source,
        last_source=last_source,
        last_actor=last_actor,
        companions=companions,
        answers=answers,
    )


def _question_meta(db: Session, wedding: Wedding) -> dict:
    """question_id → (prompt, qtype) for the wedding (to label embedded answers)."""
    return {
        q.id: (q.prompt, q.qtype.value)
        for q in db.execute(
            select(Question).where(Question.wedding_id == wedding.id)
        ).scalars()
    }


def _questions_ordered(db: Session, wedding: Wedding) -> list[Question]:
    """The wedding's questions in column/render order (one export column each)."""
    return list(
        db.execute(
            select(Question)
            .where(Question.wedding_id == wedding.id)
            .order_by(Question.sort_order, Question.prompt)
        ).scalars()
    )


# Eager-load everything _guest_admin walks (rsvp → companions → answers, plus the
# rsvp's own answers). Without this the guest list / export lazy-load per row —
# ~600+ queries for a 200-guest wedding, which over the serverless → Supabase
# pooler hop (NullPool, one round-trip each) dominates the response time.
_GUEST_ADMIN_LOADS = (
    selectinload(Guest.rsvp).selectinload(Rsvp.companions).selectinload(Companion.answers),
    selectinload(Guest.rsvp).selectinload(Rsvp.answers),
)


def _guests_with_rsvp(db: Session, wedding: Wedding) -> list[Guest]:
    return (
        db.execute(
            select(Guest)
            .where(Guest.wedding_id == wedding.id)
            .order_by(Guest.name)
            .options(*_GUEST_ADMIN_LOADS)
        )
        .scalars()
        .all()
    )


def _validated_contacts(payload, wedding: Wedding) -> dict:
    """Normalize email/phone present on a guest create/update payload (E.164 phone,
    format-checked email). National-format phones are interpreted in the wedding's
    configured region. Returns only the keys that were set."""
    out: dict = {}
    data = payload.model_dump(exclude_unset=True)
    try:
        if "email" in data:
            out["email"] = normalize_email(data["email"])
        if "phone" in data:
            out["phone"] = normalize_phone(data["phone"], region=wedding_phone_region(wedding))
    except ContactError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return out


def _question_admin(q: Question) -> QuestionAdmin:
    return QuestionAdmin(
        id=q.id,
        prompt=q.prompt,
        qtype=q.qtype.value,
        options=q.options,
        required=q.required,
        scope=q.scope.value,
        applies_to=q.applies_to.value,
        visibility=q.visibility.value,
        visibility_ref=q.visibility_ref,
        sort_order=q.sort_order,
    )


def _story_arc_admin(a: StoryArc) -> StoryArcAdmin:
    return StoryArcAdmin(
        id=a.id, title=a.title, visible=a.visible, sort_order=a.sort_order, content=a.content
    )


def _deep_merge(base: dict, patch: dict) -> dict:
    """Recursively merge `patch` onto `base` (dicts merge; everything else,
    including lists, is replaced). Returns a new dict; inputs aren't mutated."""
    out = dict(base)
    for key, value in patch.items():
        if isinstance(value, dict) and isinstance(out.get(key), dict):
            out[key] = _deep_merge(out[key], value)
        else:
            out[key] = value
    return out


# --- Whoami ----------------------------------------------------------------
def _can_publish(ctx: WeddingCtx) -> bool:
    """Publish rights: owner (and platform admin) always; a co-admin only when
    the owner granted it via the per-wedding `admins_can_publish` setting."""
    if ctx.role in ("owner", "platform"):
        return True
    return bool((ctx.wedding.settings or {}).get("admins_can_publish"))


@router.get("/me", response_model=AdminMe)
def me(ctx: WeddingCtx = Depends(member_ctx), db: Session = Depends(get_db)) -> AdminMe:
    wedding = ctx.wedding
    return AdminMe(
        email=ctx.user.email,
        via=ctx.user.via,
        wedding_id=wedding.id,
        wedding_slug=wedding.slug,
        couple_names=wedding.couple_names,
        role=ctx.role,
        wedding_status=wedding.status,
        published=wedding.published,
        can_publish=_can_publish(ctx),
        entitlements=effective_entitlements(db, wedding),
        storage_bytes_used=wedding.storage_bytes_used or 0,
    )


# --- Guests ----------------------------------------------------------------
@router.get("/guests", response_model=list[GuestAdmin])
def list_guests(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> list[GuestAdmin]:
    guests = _guests_with_rsvp(db, wedding)
    qmeta = _question_meta(db, wedding)
    return [_guest_admin(g, qmeta, wedding.content) for g in guests]


@router.post("/guests", response_model=GuestAdmin, status_code=201)
def create_guest(
    payload: GuestCreate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> GuestAdmin:
    check_limit(db, wedding, "max_guests", _guest_count(db, wedding))
    contacts = _validated_contacts(payload, wedding)
    tier = InviteTier(payload.invite_tier)
    guest = Guest(
        wedding_id=wedding.id,
        # Name is optional now; fall back to the greeting so the slug is still readable.
        slug=_unique_slug(db, payload.name or payload.greeting_name),
        name=payload.name,
        greeting_name=payload.greeting_name,
        party_members=clamp_party_members(payload.party_members, tier, wedding.content) or None,
        email=contacts.get("email"),
        phone=contacts.get("phone"),
        side=payload.side,
        relationship_label=payload.relationship,
        group_name=payload.group_name,
        batch=payload.batch,
        invite_tier=tier,
        invited=payload.invited,
        expected_party_size=payload.expected_party_size,
        story_arc_ids=_validated_arc_ids(db, wedding, payload.story_arc_ids) or None,
        seed_meta={"source": "admin"},
    )
    db.add(guest)
    db.commit()
    db.refresh(guest)
    return _guest_admin(guest, _question_meta(db, wedding), wedding.content)


def _guest_count(db: Session, wedding: Wedding) -> int:
    return db.execute(
        select(func.count()).select_from(Guest).where(Guest.wedding_id == wedding.id)
    ).scalar_one()


def _get_owned_guest(db: Session, wedding: Wedding, guest_id: UUID) -> Guest:
    guest = db.execute(
        select(Guest).where(Guest.id == guest_id, Guest.wedding_id == wedding.id)
    ).scalar_one_or_none()
    if guest is None:
        raise HTTPException(status_code=404, detail="Guest not found")
    return guest


@router.patch("/guests/{guest_id}", response_model=GuestAdmin)
def update_guest(
    guest_id: UUID,
    payload: GuestUpdate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> GuestAdmin:
    guest = _get_owned_guest(db, wedding, guest_id)
    contacts = _validated_contacts(payload, wedding)  # normalize before we touch the row
    data = payload.model_dump(exclude_unset=True)
    if "invite_tier" in data:
        guest.invite_tier = InviteTier(data.pop("invite_tier"))
    if "relationship" in data:
        guest.relationship_label = data.pop("relationship")
    for field in ("email", "phone"):
        if field in data:
            # Use the normalized value (data still holds the raw input).
            setattr(guest, field, contacts.get(field))
            data.pop(field)
    if "story_arc_ids" in data:
        # [] / null both clear the override (back to "all visible arcs"); stored
        # as NULL. Non-empty is validated to this wedding's arcs.
        ids = _validated_arc_ids(db, wedding, data.pop("story_arc_ids"))
        guest.story_arc_ids = ids or None
    if "party_members" in data:
        # Clamp the prefill party to the (possibly just-changed) tier; [] → NULL.
        guest.party_members = (
            clamp_party_members(data.pop("party_members"), guest.invite_tier, wedding.content)
            or None
        )
    for field, value in data.items():
        setattr(guest, field, value)
    db.commit()
    db.refresh(guest)
    return _guest_admin(guest, _question_meta(db, wedding), wedding.content)


@router.delete("/guests/{guest_id}", status_code=204)
def delete_guest(
    guest_id: UUID,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> Response:
    guest = _get_owned_guest(db, wedding, guest_id)
    db.delete(guest)
    db.commit()
    return Response(status_code=204)


def _set_rsvp_status(
    db: Session,
    wedding: Wedding,
    guest: Guest,
    status: str,
    *,
    notes: str | None = None,
    actor: str | None = None,
) -> Rsvp | None:
    """Apply just the attendance status to one guest (no answers). Shared by the
    single-guest RSVP override and the bulk action. Returns the live Rsvp, or None
    when `pending`/`invited` cleared it. `attending` keeps any existing party/answers
    (so a bulk "mark attending" doesn't wipe replies); `declined` drops the party.
    `pending`/`invited` both clear the RSVP back to "no reply" and set the
    `invite_sent` flag accordingly (invited = the owner has sent the invite).

    Stamps the RSVP audit trail as an admin write (`actor` = the owner's email)."""
    rsvp = guest.rsvp
    if status in ("pending", "invited"):
        # No reply yet — drop any RSVP and record whether the invite has been sent.
        guest.invite_sent = status == "invited"
        if rsvp is not None:
            db.delete(rsvp)
            guest.rsvp = None
        return None
    attending = status == "attending"
    if rsvp is None:
        rsvp = Rsvp(wedding_id=wedding.id, attending=attending)
        guest.rsvp = rsvp
        db.flush()
    rsvp.attending = attending
    if notes is not None:
        rsvp.notes = notes
    if not attending:
        # A declined RSVP has no party.
        rsvp.companions[:] = []
        rsvp.answers[:] = []
    elif not rsvp.companions:
        # Newly attending with no party yet: materialize the admin's prefill +1/kids
        # (party_members) into real Companion rows so their per-person answers (age,
        # dietary, …) become editable in the by-person view. Only when the party is
        # empty, so a bulk "mark attending" never clobbers an existing/responded party.
        rsvp.companions[:] = [
            Companion(wedding_id=wedding.id, kind=CompanionKind(m["kind"]), name=m["name"] or None)
            for m in clamp_party_members(guest.party_members, guest.invite_tier, wedding.content)
        ]
    stamp_rsvp(rsvp, SOURCE_ADMIN, actor=actor)
    return rsvp


@router.put("/guests/{guest_id}/rsvp", response_model=GuestAdmin)
def set_guest_rsvp(
    guest_id: UUID,
    payload: GuestRsvpUpdate,
    ctx: WeddingCtx = Depends(editor_ctx),
    db: Session = Depends(get_db),
) -> GuestAdmin:
    wedding = ctx.wedding
    owner = ctx.user
    """Owner override of a guest's RSVP (status + the whole party). Lets the couple
    record a response on a guest's behalf or correct one. `pending` removes the RSVP;
    `declined` keeps it but clears the party; `attending` sets the primary's own +
    invitee-scope `answers` and — when `companions` is given — the entire +1/child
    party with each person's own answers (wholesale replace, like the guest's own
    submit). Omitting `companions` leaves the companion party untouched."""
    guest = _get_owned_guest(db, wedding, guest_id)
    rsvp = _set_rsvp_status(
        db, wedding, guest, payload.status, notes=payload.notes, actor=owner.email
    )

    if rsvp is not None and payload.status == "attending":
        questions = {
            q.id: q
            for q in db.execute(
                select(Question).where(Question.wedding_id == wedding.id)
            ).scalars()
        }
        if payload.answers is not None:
            for a in payload.answers:
                q = questions.get(a.question_id)
                if q is None or not is_party_question(q):
                    raise HTTPException(status_code=422, detail="Question not applicable")
            # Replace only the party answers (companion_id NULL); companion ones below.
            for a in list(rsvp.answers):
                if a.companion_id is None:
                    rsvp.answers.remove(a)
            rsvp.answers.extend(
                Answer(
                    wedding_id=wedding.id,
                    question_id=a.question_id,
                    value=a.value,
                    companion_id=None,
                )
                for a in payload.answers
            )
        if payload.companions is not None:
            # Wholesale replace the +1/child party (+ each person's answers), gated by
            # the tier's caps and each question's scope — mirrors the guest submit and
            # the import path. Old companions (and their answers) cascade away.
            caps = capabilities_for(guest.invite_tier, wedding.content)
            adults = [c for c in payload.companions if c.kind == "adult"]
            children = [c for c in payload.companions if c.kind == "child"]
            if (
                len(adults) > caps.max_adult_companions
                or len(children) > caps.max_child_companions
            ):
                raise HTTPException(
                    status_code=422, detail="Too many companions for this invitation"
                )
            new_companions = [
                Companion(wedding_id=wedding.id, kind=CompanionKind(c.kind), name=c.name)
                for c in payload.companions
            ]
            rsvp.companions[:] = new_companions
            db.flush()  # assign companion ids before attaching their answers
            for comp, c in zip(new_companions, payload.companions):
                is_child = comp.kind is CompanionKind.child
                for a in c.answers:
                    q = questions.get(a.question_id)
                    if q is None or not person_question_applies(q, is_child=is_child):
                        raise HTTPException(status_code=422, detail="Question not applicable")
                    rsvp.answers.append(
                        Answer(
                            wedding_id=wedding.id,
                            question_id=a.question_id,
                            value=a.value,
                            companion_id=comp.id,
                        )
                    )

    db.commit()
    db.refresh(guest)
    return _guest_admin(guest, _question_meta(db, wedding), wedding.content)


# --- Bulk guest actions ----------------------------------------------------
def _owned_guests_in(db: Session, wedding: Wedding, ids: list[UUID]) -> list[Guest]:
    """The subset of `ids` that are guests of THIS wedding (tenant guard). Foreign
    or unknown ids simply don't come back, so a bulk action can never reach across
    tenants — mirrors the per-guest `_get_owned_guest` 404."""
    return list(
        db.execute(
            select(Guest).where(Guest.id.in_(ids), Guest.wedding_id == wedding.id)
        )
        .scalars()
        .all()
    )


@router.post("/guests/bulk/rsvp", response_model=BulkResult)
def bulk_set_rsvp(
    payload: BulkRsvpUpdate,
    ctx: WeddingCtx = Depends(editor_ctx),
    db: Session = Depends(get_db),
) -> BulkResult:
    """Set the attendance status for many guests at once (the only bulk-editable
    field besides delete). Answers are left untouched; see `_set_rsvp_status`."""
    wedding, owner = ctx.wedding, ctx.user
    guests = _owned_guests_in(db, wedding, payload.ids)
    for g in guests:
        _set_rsvp_status(db, wedding, g, payload.status, actor=owner.email)
    db.commit()
    return BulkResult(count=len(guests))


@router.post("/guests/bulk/delete", response_model=BulkResult)
def bulk_delete_guests(
    payload: BulkGuestIds,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> BulkResult:
    """Remove many guests at once (each cascades to its RSVP, like the single
    delete)."""
    guests = _owned_guests_in(db, wedding, payload.ids)
    for g in guests:
        db.delete(g)
    db.commit()
    return BulkResult(count=len(guests))


# --- Companions (the +1 / kids on an RSVP) ---------------------------------
def _get_owned_companion(db: Session, wedding: Wedding, companion_id: UUID) -> Companion:
    c = db.execute(
        select(Companion).where(
            Companion.id == companion_id, Companion.wedding_id == wedding.id
        )
    ).scalar_one_or_none()
    if c is None:
        raise HTTPException(status_code=404, detail="Companion not found")
    return c


@router.patch("/companions/{companion_id}", response_model=CompanionAdmin)
def update_companion(
    companion_id: UUID,
    payload: CompanionUpdate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> CompanionAdmin:
    """Edit a single companion's name and/or its person-scope answers. `kind` stays
    as authored (adult vs child is structural and tier-gated). When `answers` is
    given it replaces this companion's answers wholesale, validated against the
    wedding's questions that apply to this person's kind."""
    companion = _get_owned_companion(db, wedding, companion_id)
    data = payload.model_dump(exclude_unset=True)
    if "name" in data:
        companion.name = data["name"]
    if payload.answers is not None:
        is_child = companion.kind is CompanionKind.child
        questions = {
            q.id: q
            for q in db.execute(
                select(Question).where(Question.wedding_id == wedding.id)
            ).scalars()
        }
        for a in payload.answers:
            q = questions.get(a.question_id)
            if q is None or not person_question_applies(q, is_child=is_child):
                raise HTTPException(status_code=422, detail="Question not applicable")
        companion.answers[:] = [
            Answer(
                wedding_id=wedding.id,
                rsvp_id=companion.rsvp_id,
                question_id=a.question_id,
                value=a.value,
                companion_id=companion.id,
            )
            for a in payload.answers
        ]
    db.commit()
    db.refresh(companion)
    return CompanionAdmin(
        id=companion.id,
        kind=companion.kind.value,
        name=companion.name,
        answers=_answer_admins(companion.answers, _question_meta(db, wedding)),
    )


@router.delete("/companions/{companion_id}", status_code=204)
def delete_companion(
    companion_id: UUID,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> Response:
    companion = _get_owned_companion(db, wedding, companion_id)
    db.delete(companion)
    db.commit()
    return Response(status_code=204)


# --- Questions -------------------------------------------------------------
@router.get("/questions", response_model=list[QuestionAdmin])
def list_questions(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> list[QuestionAdmin]:
    rows = (
        db.execute(
            select(Question)
            .where(Question.wedding_id == wedding.id)
            .order_by(Question.sort_order, Question.prompt)
        )
        .scalars()
        .all()
    )
    return [_question_admin(q) for q in rows]


@router.post("/questions", response_model=QuestionAdmin, status_code=201)
def create_question(
    payload: QuestionCreate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> QuestionAdmin:
    check_limit(
        db, wedding, "max_custom_questions",
        db.execute(
            select(func.count()).select_from(Question).where(Question.wedding_id == wedding.id)
        ).scalar_one(),
    )
    q = Question(
        wedding_id=wedding.id,
        prompt=payload.prompt,
        qtype=QuestionType(payload.qtype),
        options=payload.options,
        required=payload.required,
        scope=QuestionScope(payload.scope),
        applies_to=QuestionApplies(payload.applies_to),
        visibility=QuestionVisibility(payload.visibility),
        visibility_ref=payload.visibility_ref,
        sort_order=payload.sort_order,
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    return _question_admin(q)


def _get_owned_question(db: Session, wedding: Wedding, question_id: UUID) -> Question:
    q = db.execute(
        select(Question).where(
            Question.id == question_id, Question.wedding_id == wedding.id
        )
    ).scalar_one_or_none()
    if q is None:
        raise HTTPException(status_code=404, detail="Question not found")
    return q


@router.patch("/questions/{question_id}", response_model=QuestionAdmin)
def update_question(
    question_id: UUID,
    payload: QuestionUpdate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> QuestionAdmin:
    q = _get_owned_question(db, wedding, question_id)
    data = payload.model_dump(exclude_unset=True)
    if "qtype" in data:
        q.qtype = QuestionType(data.pop("qtype"))
    if "visibility" in data:
        q.visibility = QuestionVisibility(data.pop("visibility"))
    if "scope" in data:
        q.scope = QuestionScope(data.pop("scope"))
    if "applies_to" in data:
        q.applies_to = QuestionApplies(data.pop("applies_to"))
    for field, value in data.items():
        setattr(q, field, value)
    db.commit()
    db.refresh(q)
    return _question_admin(q)


@router.delete("/questions/{question_id}", status_code=204)
def delete_question(
    question_id: UUID,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> Response:
    q = _get_owned_question(db, wedding, question_id)
    db.delete(q)
    db.commit()
    return Response(status_code=204)


# --- Content / event details / theme ---------------------------------------
@router.get("/content", response_model=ContentAdmin)
def get_content(wedding: Wedding = Depends(member_wedding)) -> ContentAdmin:
    return ContentAdmin(
        couple_names=wedding.couple_names,
        event_details=wedding.event_details or {},
        content=wedding.content or {},
        theme_tokens=wedding.theme_tokens,
    )


@router.patch("/content", response_model=ContentAdmin)
def update_content(
    payload: ContentUpdate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> ContentAdmin:
    """Partial edit of the invite's copy/details/theme. JSON sections deep-merge so
    the admin can PATCH a single block (e.g. just `content.cover`); `couple_names`
    is replaced. JSON columns are reassigned (not mutated) so SQLAlchemy flushes."""
    data = payload.model_dump(exclude_unset=True)
    if "couple_names" in data and data["couple_names"] is not None:
        wedding.couple_names = data["couple_names"]
    if data.get("event_details") is not None:
        wedding.event_details = _deep_merge(wedding.event_details or {}, data["event_details"])
    if data.get("content") is not None:
        wedding.content = _deep_merge(wedding.content or {}, data["content"])
    if "theme_tokens" in data:
        tokens = data["theme_tokens"]
        wedding.theme_tokens = (
            _deep_merge(wedding.theme_tokens or {}, tokens) if tokens is not None else None
        )
    db.commit()
    db.refresh(wedding)
    return get_content(wedding)


# --- Story arcs ------------------------------------------------------------
@router.get("/story-arcs", response_model=list[StoryArcAdmin])
def list_story_arcs(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> list[StoryArcAdmin]:
    rows = (
        db.execute(
            select(StoryArc)
            .where(StoryArc.wedding_id == wedding.id)
            .order_by(StoryArc.sort_order, StoryArc.title)
        )
        .scalars()
        .all()
    )
    return [_story_arc_admin(a) for a in rows]


@router.post("/story-arcs", response_model=StoryArcAdmin, status_code=201)
def create_story_arc(
    payload: StoryArcCreate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> StoryArcAdmin:
    check_limit(
        db, wedding, "max_story_arcs",
        db.execute(
            select(func.count()).select_from(StoryArc).where(StoryArc.wedding_id == wedding.id)
        ).scalar_one(),
    )
    arc = StoryArc(
        wedding_id=wedding.id,
        title=payload.title,
        visible=payload.visible,
        sort_order=payload.sort_order,
        content=payload.content,
    )
    db.add(arc)
    db.commit()
    db.refresh(arc)
    return _story_arc_admin(arc)


def _get_owned_arc(db: Session, wedding: Wedding, arc_id: UUID) -> StoryArc:
    arc = db.execute(
        select(StoryArc).where(StoryArc.id == arc_id, StoryArc.wedding_id == wedding.id)
    ).scalar_one_or_none()
    if arc is None:
        raise HTTPException(status_code=404, detail="Story arc not found")
    return arc


@router.patch("/story-arcs/{arc_id}", response_model=StoryArcAdmin)
def update_story_arc(
    arc_id: UUID,
    payload: StoryArcUpdate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> StoryArcAdmin:
    arc = _get_owned_arc(db, wedding, arc_id)
    data = payload.model_dump(exclude_unset=True)
    # `content` is replaced wholesale (the editor sends the full arc body).
    for field, value in data.items():
        setattr(arc, field, value)
    db.commit()
    db.refresh(arc)
    return _story_arc_admin(arc)


@router.delete("/story-arcs/{arc_id}", status_code=204)
def delete_story_arc(
    arc_id: UUID,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> Response:
    arc = _get_owned_arc(db, wedding, arc_id)
    db.delete(arc)
    db.commit()
    return Response(status_code=204)


# --- Image upload (story beats / iconography) ------------------------------
@router.post("/upload", response_model=UploadResult)
async def upload_image(
    file: UploadFile = File(...),
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> UploadResult:
    """Accept a single image (multipart) and return a stored URL to drop into a
    story beat's `image`. Namespaced per wedding; dev → local disk, prod →
    Supabase Storage (see app/storage.py). Gated by the plan's `max_storage_mb`
    against the wedding's byte counter (compressed size, checked pre-persist)."""
    data = await file.read()
    try:
        blob, ext = prepare_image(data, file.content_type)
    except UploadError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    check_storage(db, wedding, adding_bytes=len(blob))
    try:
        url = store_image(settings, wedding.slug, blob, ext, file.content_type or "image/png")
    except UploadError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    wedding.storage_bytes_used = (wedding.storage_bytes_used or 0) + len(blob)
    db.commit()
    return UploadResult(url=url)


# --- Responses + summary ---------------------------------------------------
def _attending_rsvps(db: Session, wedding: Wedding) -> list[Rsvp]:
    return (
        db.execute(
            select(Rsvp)
            .where(Rsvp.wedding_id == wedding.id)
            # Callers walk r.companions and r.answers per row — eager-load both so
            # summary/timeline don't fan out into one query per RSVP against Supabase.
            .options(selectinload(Rsvp.companions), selectinload(Rsvp.answers))
        )
        .scalars()
        .all()
    )


@router.get("/responses", response_model=list[ResponseAdmin])
def list_responses(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> list[ResponseAdmin]:
    rsvps = (
        db.execute(
            select(Rsvp)
            .where(Rsvp.wedding_id == wedding.id)
            # Most-recently-changed first, so new RSVPs and edits surface at the top.
            .order_by(Rsvp.updated_at.desc())
            # Serializer walks guest + companions(+answers) + answers per row.
            .options(
                selectinload(Rsvp.guest),
                selectinload(Rsvp.companions).selectinload(Companion.answers),
                selectinload(Rsvp.answers),
            )
        )
        .scalars()
        .all()
    )
    qmeta = _question_meta(db, wedding)
    out: list[ResponseAdmin] = []
    for r in rsvps:
        out.append(
            ResponseAdmin(
                guest_id=r.guest_id,
                guest_name=r.guest.name,
                slug=r.guest.slug,
                attending=r.attending,
                responded_at=r.responded_at,
                updated_at=r.updated_at,
                last_source=r.last_source,
                last_actor=r.last_actor,
                notes=r.notes,
                companions=[
                    CompanionAdmin(
                        id=c.id,
                        kind=c.kind.value,
                        name=c.name,
                        answers=_answer_admins(c.answers, qmeta),
                    )
                    for c in r.companions
                ],
                answers=_answer_admins([a for a in r.answers if a.companion_id is None], qmeta),
            )
        )
    return out


# Question types that produce a categorical tally (the rest — text/number — aren't
# charted, though a `number` like age could be bucketed later).
_CATEGORICAL_QTYPES = {"choice", "multi_choice", "yesno"}


def _naive_utc(dt: datetime) -> datetime:
    """Normalize a stored timestamp to naive-UTC for arithmetic. Postgres returns
    tz-aware datetimes; SQLite (tests/local) returns naive — coerce both so window
    comparisons never mix aware/naive."""
    return dt if dt.tzinfo is None else dt.astimezone(timezone.utc).replace(tzinfo=None)


def _reply_times(rsvps: list[Rsvp]) -> list[datetime]:
    """First-reply timestamps (naive UTC), oldest first. One per RSVP; edits don't
    add points (we use responded_at, not updated_at)."""
    return sorted(_naive_utc(r.responded_at) for r in rsvps if r.responded_at)


def _expected_for_guest(guest: Guest, content: dict | None = None) -> int:
    """Best pre-RSVP estimate of how many people this invite brings, by the owner's
    chosen fallback chain: the explicit `expected_party_size` (already incl. the
    invitee) → else 1 + the prefilled `party_members` count → else a conservative
    floor of the invitee + one likely extra adult when companions are allowed
    (optional kids and the *additional* family adults aren't assumed — the cap is a
    ceiling, not an expectation). Always ≥ 1 (the invitee). `content` threads the
    wedding's configured plus_family caps."""
    if guest.expected_party_size is not None:
        return max(guest.expected_party_size, 0)
    members = clamp_party_members(guest.party_members, guest.invite_tier, content)
    if members:
        return 1 + len(members)
    allows_adult = capabilities_for(guest.invite_tier, content).max_adult_companions > 0
    return 1 + (1 if allows_adult else 0)


def _answer_labels(value: dict, qtype: str) -> list[str]:
    """The categorical label(s) an answer contributes to a breakdown. multi_choice
    yields one per selected option; choice/yesno yield one; text/number yield none."""
    if not isinstance(value, dict):
        return []
    if qtype == "choice":
        v = value.get("choice")
        return [str(v)] if v not in (None, "") else []
    if qtype == "yesno":
        v = value.get("yesno")
        return [] if v is None else ["Yes" if v else "No"]
    if qtype == "multi_choice":
        return [str(x) for x in (value.get("choices") or []) if x not in (None, "")]
    return []


def _order_counts(q: Question, counts: dict[str, int]) -> list[OptionCount]:
    """Order a tally by the question's own option order (so charts read predictably),
    with any unexpected values appended by descending count."""
    declared = ["Yes", "No"] if q.qtype.value == "yesno" else [str(o) for o in (q.options or [])]
    ordered: list[OptionCount] = []
    seen: set[str] = set()
    for label in declared:
        if label in counts:
            ordered.append(OptionCount(label=label, count=counts[label]))
            seen.add(label)
    extras = sorted(
        ((k, v) for k, v in counts.items() if k not in seen), key=lambda kv: (-kv[1], kv[0])
    )
    ordered += [OptionCount(label=k, count=v) for k, v in extras]
    return ordered


def _question_breakdowns(
    questions: list[Question], rsvps: list[Rsvp]
) -> list[QuestionBreakdown]:
    """Tally every categorical question across ATTENDING people/parties. Generic over
    the question engine — `applicable` counts who could answer (by scope/applies_to),
    `answered` who did, and `counts` the per-value bars."""
    out: list[QuestionBreakdown] = []
    for q in questions:
        if q.qtype.value not in _CATEGORICAL_QTYPES:
            continue
        applicable = answered = 0
        counts: dict[str, int] = {}
        for r in rsvps:
            if not r.attending:
                continue
            # Which "persons" in this party are asked q? None = the primary/party row.
            persons: list = []
            if is_party_question(q):  # invitee-scope, or person-scope hitting the primary
                persons.append(None)
            if q.scope is QuestionScope.person:
                persons += [
                    c.id
                    for c in r.companions
                    if person_question_applies(q, is_child=(c.kind is CompanionKind.child))
                ]
            ans = {a.companion_id: a.value for a in r.answers if a.question_id == q.id}
            for pid in persons:
                applicable += 1
                labels = _answer_labels(ans.get(pid) or {}, q.qtype.value)
                if labels:
                    answered += 1
                    for label in labels:
                        counts[label] = counts.get(label, 0) + 1
        out.append(
            QuestionBreakdown(
                question_id=q.id,
                prompt=q.prompt,
                qtype=q.qtype.value,
                scope=q.scope.value,
                applies_to=q.applies_to.value,
                applicable=applicable,
                answered=answered,
                counts=_order_counts(q, counts),
            )
        )
    return out


# The dimensions the Overview pivot can group/stack by, in selector order. `status`
# is the four-bucket invitation status; `tier` is the enum; the rest are free-text
# guest columns ("" → Unassigned). The same dimension can drive either axis
# (Group by = bars/rows, Then by = the stacked segments).
_PIVOT_DIMS = ("status", "side", "batch", "tier", "relationship", "group")
_TIER_LABELS = {"solo": "Solo", "plus_one": "Plus-one", "plus_family": "Family"}
_STATUS_LABELS = {"attending": "Attending", "declined": "Declined", "invited": "Invited", "pending": "Pending"}
# Funnel order for the status dimension (so a status-stack always reads the same way).
_STATUS_ORDER = {"attending": 0, "declined": 1, "invited": 2, "pending": 3}
_UNASSIGNED = "__unassigned__"


def _guest_status(g: Guest) -> str:
    """The guest's invitation status — one of pending/invited/attending/declined.
    Mirrors the per-row status the dashboard shows: no RSVP yet splits into `invited`
    (owner sent the link) vs `pending` (not yet contacted)."""
    r = g.rsvp
    if r is None:
        return "invited" if g.invite_sent else "pending"
    return "attending" if r.attending else "declined"


def _guest_people(g: Guest, content: dict | None) -> int:
    """How many PEOPLE this invite represents for the headcount lens: the real
    confirmed party if they're attending, otherwise the owner's expected estimate
    (so Invited/Pending/Declined are counted at their estimated size)."""
    if _guest_status(g) == "attending" and g.rsvp is not None:
        return 1 + len(g.rsvp.companions)
    return _expected_for_guest(g, content)


def _dim_bucket(g: Guest, dim: str) -> tuple[str, str]:
    """(key, label) of the bucket guest `g` falls into for dimension `dim`. The
    no-value bucket is keyed `__unassigned__` and labelled "Unassigned"."""
    if dim == "status":
        s = _guest_status(g)
        return s, _STATUS_LABELS[s]
    if dim == "tier":
        v = g.invite_tier.value
        return v, _TIER_LABELS.get(v, v)
    raw = {
        "side": g.side,
        "batch": g.batch,
        "relationship": g.relationship_label,
        "group": g.group_name,
    }.get(dim)
    key = (raw or "").strip()
    return (key, key) if key else (_UNASSIGNED, "Unassigned")


def _available_dims(guests: list[Guest]) -> list[str]:
    """Dimensions worth offering in the pivot selectors — those that actually split the
    list into ≥2 buckets with at least one real (non-Unassigned) value. Keeps the
    menu free of dimensions nobody has filled in (e.g. no batches yet)."""
    out: list[str] = []
    for dim in _PIVOT_DIMS:
        keys = {_dim_bucket(g, dim)[0] for g in guests}
        if len(keys) >= 2 and any(k != _UNASSIGNED for k in keys):
            out.append(dim)
    return out


def _accumulate(guests: list[Guest], content: dict | None) -> dict:
    """Roll a set of guests up into the GroupBreakdown count fields: the status counts
    (invitation lens), confirmed/expected headcount, and the mixed `people` measure."""
    cell = {
        "invitations": 0, "pending": 0, "invited": 0, "attending": 0,
        "declined": 0, "head_count": 0, "invited_people": 0,
        "expected_head_count": 0, "people": 0,
    }
    for g in guests:
        cell["invitations"] += 1
        cell["expected_head_count"] += _expected_for_guest(g, content)
        cell["people"] += _guest_people(g, content)
        status = _guest_status(g)
        cell[status] += 1
        if status == "attending" and g.rsvp is not None:
            cell["head_count"] += 1 + len(g.rsvp.companions)
        elif status == "invited":
            # Capacity lens: people we've committed to but who haven't replied yet.
            cell["invited_people"] += _expected_for_guest(g, content)
    return cell


def _capacity_config(event_details: dict | None) -> CapacityConfig:
    """Parse the owner's capacity (people) out of `event_details.capacity`, tolerating
    a missing/partial/garbage blob (it's free-form JSON the admin edits). Non-numeric
    or negative values are dropped rather than raising."""
    raw = (event_details or {}).get("capacity") or {}
    if not isinstance(raw, dict):
        return CapacityConfig()

    def _as_cap(v: object) -> int | None:
        # bool is an int subclass — exclude it so a stray `true` isn't read as 1.
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            return None
        return int(v) if v >= 0 else None

    total = _as_cap(raw.get("total"))
    by_side_raw = raw.get("by_side")
    by_side: dict[str, int] = {}
    if isinstance(by_side_raw, dict):
        for label, value in by_side_raw.items():
            cap = _as_cap(value)
            if cap is not None and str(label).strip():
                by_side[str(label)] = cap
    return CapacityConfig(total=total, by_side=by_side)


def _sort_groups(rows: list[GroupBreakdown], dim: str) -> list[GroupBreakdown]:
    """Order a dimension's buckets. `status` follows the fixed funnel order; everything
    else is invitation count desc with "Unassigned" pinned last."""
    if dim == "status":
        return sorted(rows, key=lambda r: _STATUS_ORDER.get(r.key, 99))
    return sorted(rows, key=lambda r: (r.key == _UNASSIGNED, -r.invitations, r.label))


def _group_breakdown(
    guests: list[Guest], dim: str, content: dict | None, then: str | None = None
) -> list[GroupBreakdown]:
    """Group `guests` by `dim` into status + headcount + people rollups, optionally
    sub-grouped by `then` (one level — the stacked segments). Pure bucketing: callers
    decide whether a dimension is worth showing (see `_available_dims`)."""
    buckets: dict[str, dict] = {}
    for g in guests:
        key, label = _dim_bucket(g, dim)
        bucket = buckets.setdefault(key, {"label": label, "guests": []})
        bucket["guests"].append(g)
    rows = [
        GroupBreakdown(
            key=key,
            label=bucket["label"],
            children=_group_breakdown(bucket["guests"], then, content) if then else [],
            **_accumulate(bucket["guests"], content),
        )
        for key, bucket in buckets.items()
    ]
    return _sort_groups(rows, dim)


@router.get("/summary", response_model=AdminSummary)
def summary(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> AdminSummary:
    guests = list(
        db.execute(
            select(Guest)
            .where(Guest.wedding_id == wedding.id)
            # Same N+1 guard as the pivot: _guest_status/_group_breakdown walk
            # g.rsvp(.companions) per guest.
            .options(selectinload(Guest.rsvp).selectinload(Rsvp.companions))
        ).scalars()
    )
    total_guests = len(guests)
    # Owner's planning estimate across the whole list (fallback chain per guest).
    expected_head_count = sum(_expected_for_guest(g, wedding.content) for g in guests)
    rsvps = _attending_rsvps(db, wedding)

    attending = sum(1 for r in rsvps if r.attending)
    declined = sum(1 for r in rsvps if not r.attending)
    # Guests with no reply split into "invited" (owner sent the invite) and "pending".
    invited = sum(1 for g in guests if g.rsvp is None and g.invite_sent)
    pending = total_guests - attending - declined - invited

    head_count = 0
    extra_adults = 0
    extra_children = 0

    for r in rsvps:
        if not r.attending:
            continue
        head_count += 1 + len(r.companions)
        for c in r.companions:
            if c.kind.value == "adult":
                extra_adults += 1
            else:
                extra_children += 1

    breakdowns = _question_breakdowns(_questions_ordered(db, wedding), rsvps)

    # PEOPLE per status for the hero's people-lens bar: attending = confirmed heads,
    # the rest = the owner's expected estimate (so "still to chase" is in people).
    status_people = {"attending": 0, "invited": 0, "pending": 0, "declined": 0}
    for g in guests:
        status_people[_guest_status(g)] += _guest_people(g, wedding.content)

    # Momentum: new replies in the trailing 7 days vs the 7 before (by first reply).
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    wk1, wk2 = now - timedelta(days=7), now - timedelta(days=14)
    times = _reply_times(rsvps)
    replies_this_week = sum(1 for t in times if wk1 < t <= now)
    replies_last_week = sum(1 for t in times if wk2 < t <= wk1)

    dims = _available_dims(guests)
    return AdminSummary(
        total_guests=total_guests,
        attending=attending,
        declined=declined,
        invited=invited,
        pending=max(pending, 0),
        invite_sent_count=sum(1 for g in guests if g.invite_sent),
        head_count=head_count,
        invited_people=status_people["invited"],
        pending_people=status_people["pending"],
        declined_people=status_people["declined"],
        extra_adults=extra_adults,
        extra_children=extra_children,
        expected_head_count=int(expected_head_count),
        capacity=_capacity_config(wedding.event_details),
        question_breakdowns=breakdowns,
        by_side=_group_breakdown(guests, "side", wedding.content) if "side" in dims else [],
        by_tier=_group_breakdown(guests, "tier", wedding.content) if "tier" in dims else [],
        replies_this_week=replies_this_week,
        replies_last_week=replies_last_week,
    )


@router.get("/summary/pivot", response_model=PivotSummary)
def summary_pivot(
    by: str = Query("side"),
    then: str | None = Query("status"),
    side: str | None = Query(None),
    status: str | None = Query(None),
    wedding: Wedding = Depends(member_wedding),
    db: Session = Depends(get_db),
) -> PivotSummary:
    """Configurable Overview pivot. `by` chooses the bars, `then` the stacked segments
    (one level). Optional scoping filters: `side` (a side label, e.g. "Alex"/"Sam"/
    "Unassigned") and `status` (e.g. "attending" for the Confirmed tab) restrict which
    guests are counted. `available_dims` is computed from the FULL list so the selector
    menu stays stable as you filter. Both an invalid/duplicate `then` (→ single bars)
    and an unavailable `by` (→ "side", else first available) fall back. Each cell
    carries BOTH measures (invitations + people) so the frontend renders any tab off
    one response."""
    # Eager-load rsvp + companions: the bucketing below walks g.rsvp(.companions)
    # for every guest, which is fine on local SQLite but is an N+1 round-trip storm
    # against Supabase from the serverless function. selectinload collapses it to a
    # constant 3 queries regardless of guest count (guests, rsvps, companions).
    all_guests = list(
        db.execute(
            select(Guest)
            .where(Guest.wedding_id == wedding.id)
            .options(selectinload(Guest.rsvp).selectinload(Rsvp.companions))
        ).scalars()
    )
    dims = _available_dims(all_guests)
    if by not in dims:
        by = "side" if "side" in dims else (dims[0] if dims else "status")
    if then == by or then not in dims:
        then = None

    # Scope to the chosen side / status (the chart's data; the menu stays full-list).
    guests = all_guests
    if side:
        guests = [g for g in guests if _dim_bucket(g, "side")[1] == side]
    if status:
        guests = [g for g in guests if _guest_status(g) == status]

    return PivotSummary(
        by=by,
        then=then,
        available_dims=dims,
        groups=_group_breakdown(guests, by, wedding.content, then),
        total=GroupBreakdown(
            key="__total__", label="Total", children=[],
            **_accumulate(guests, wedding.content),
        ),
    )


@router.get("/summary/timeline", response_model=TimelineSummary)
def summary_timeline(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> TimelineSummary:
    """Cumulative RSVP replies bucketed by week (Monday-anchored), for the Overview's
    Trends chart. Counted by first reply (`responded_at`); `total_invitations` is the
    ceiling. Lazy-loaded by the UI only when the Trends panel is opened."""
    total_inv = db.query(Guest).filter(Guest.wedding_id == wedding.id).count()
    times = _reply_times(_attending_rsvps(db, wedding))
    points: list[TimelinePoint] = []
    if times:
        first = times[0].date()
        start = first - timedelta(days=first.weekday())  # Monday of the first reply's week
        today = datetime.now(timezone.utc).date()
        end = today - timedelta(days=today.weekday())  # Monday of the current week
        cumulative = 0
        week = start
        while week <= end:
            nxt = week + timedelta(days=7)
            new = sum(1 for t in times if week <= t.date() < nxt)
            cumulative += new
            points.append(TimelinePoint(week_start=week, new=new, cumulative=cumulative))
            week = nxt
    return TimelineSummary(total_invitations=total_inv, total_replied=len(times), points=points)


# --- Wishes / guestbook moderation -----------------------------------------
def _wish_admin(w: Wish) -> WishAdmin:
    return WishAdmin(
        id=w.id,
        name=w.name,
        message=w.message,
        approved=w.approved,
        guest_name=w.guest.name if w.guest is not None else None,
        created_at=w.created_at,
    )


@router.get("/wishes", response_model=list[WishAdmin])
def list_wishes(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> list[WishAdmin]:
    rows = (
        db.execute(
            select(Wish)
            .where(Wish.wedding_id == wedding.id)
            .order_by(Wish.created_at.desc())
            # _wish_admin reads w.guest.name — avoid one query per wish.
            .options(selectinload(Wish.guest))
        )
        .scalars()
        .all()
    )
    return [_wish_admin(w) for w in rows]


def _get_owned_wish(db: Session, wedding: Wedding, wish_id: UUID) -> Wish:
    w = db.execute(
        select(Wish).where(Wish.id == wish_id, Wish.wedding_id == wedding.id)
    ).scalar_one_or_none()
    if w is None:
        raise HTTPException(status_code=404, detail="Wish not found")
    return w


@router.patch("/wishes/{wish_id}", response_model=WishAdmin)
def moderate_wish(
    wish_id: UUID,
    payload: WishModerate,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> WishAdmin:
    wish = _get_owned_wish(db, wedding, wish_id)
    wish.approved = payload.approved
    db.commit()
    db.refresh(wish)
    return _wish_admin(wish)


@router.delete("/wishes/{wish_id}", status_code=204)
def delete_wish(
    wish_id: UUID,
    wedding: Wedding = Depends(editable_wedding),
    db: Session = Depends(get_db),
) -> Response:
    wish = _get_owned_wish(db, wedding, wish_id)
    db.delete(wish)
    db.commit()
    return Response(status_code=204)


# --- Export / template / import (split-row schema; see app/export_import.py) ---
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx(
    header: list[str],
    rows: list[dict],
    filename: str,
    list_specs: dict[str, export_import.ListSpec] | None = None,
) -> Response:
    """Render the split-row workbook: bold/frozen header, autofilter, sized columns,
    and (from `list_specs`) per-column dropdowns backed by a hidden `Lists` sheet."""
    list_specs = list_specs or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guests"

    ws.append(header)
    header_font = Font(bold=True, color="3B2F23")
    header_fill = PatternFill("solid", fgColor="EADBC8")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        # escape_formula neutralizes CSV/formula injection from guest-supplied cells
        # (names, notes, free-text answers) when the owner opens the workbook in Excel.
        ws.append([export_import.escape_formula(row.get(col, "")) for col in header])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
    for idx, col in enumerate(header, start=1):
        widest = max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(widest + 2, 10), 40)

    if list_specs:
        # Each list lives in its own column on a hidden sheet; the validation refs it
        # (handles long lists / values with commas that inline lists can't).
        lists = wb.create_sheet("Lists")
        lists.sheet_state = "hidden"
        last_row = max(len(rows) + 1, 1000)  # cover blank rows the owner adds later
        list_col = 0
        for col_no, col_name in enumerate(header, start=1):
            spec = list_specs.get(col_name)
            if not spec or not spec.values:
                continue
            list_col += 1
            lcol = get_column_letter(list_col)
            for i, val in enumerate(spec.values, start=1):
                lists[f"{lcol}{i}"] = val
            dv = DataValidation(
                type="list",
                formula1=f"Lists!${lcol}$1:${lcol}${len(spec.values)}",
                allow_blank=True,
                showErrorMessage=spec.strict,  # non-strict = suggest but allow other input
            )
            ws.add_data_validation(dv)
            tcol = get_column_letter(col_no)
            dv.add(f"{tcol}2:{tcol}{last_row}")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.xlsx")
def export_xlsx(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> Response:
    """Split-row guest + RSVP workbook (one row per person), with dropdowns."""
    require_feature(db, wedding, "export_enabled")
    questions = _questions_ordered(db, wedding)
    header = export_import.columns([q.prompt for q in questions])
    rows = export_import.build_rows(_guests_with_rsvp(db, wedding), questions)
    return _xlsx(header, rows, f"{wedding.slug}-guests.xlsx", export_import.list_columns(questions))


@router.get("/template.xlsx")
def template_xlsx(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> Response:
    """A fillable template: the split-row header + a few example rows, with dropdowns."""
    require_feature(db, wedding, "import_enabled")
    questions = _questions_ordered(db, wedding)
    header = export_import.columns([q.prompt for q in questions])
    return _xlsx(
        header, export_import.template_rows(), f"{wedding.slug}-template.xlsx",
        export_import.list_columns(questions),
    )


def _read_records(filename: str | None, data: bytes) -> list[dict]:
    """Parse an uploaded CSV/XLSX into header-keyed dicts, each tagged with its
    1-based spreadsheet row (`__row__`) for error messages."""
    name = (filename or "").lower()
    rows: list[dict] = []
    if name.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
        for i, rec in enumerate(reader):
            rec = {(k or "").strip(): v for k, v in rec.items()}
            rec["__row__"] = i + 2
            rows.append(rec)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            return []
        for i, raw in enumerate(it):
            rec = {header[j]: raw[j] for j in range(min(len(header), len(raw)))}
            rec["__row__"] = i + 2
            rows.append(rec)
    return rows


def _resolve_import_answers(pg, qmap: dict[str, Question]):
    """Type + filter a parsed guest's answer cells against the wedding's questions.

    Returns `(party_vals, comp_vals, errors)` where `party_vals` is a list of
    `(Question, value)` for the primary (party questions only) and `comp_vals` is a
    list aligned to `pg.companions`, each a `(Question, value)` list. Unknown columns
    and answers that don't apply to a person are ignored; bad values are errors."""

    def resolve(cells: dict[str, str], *, is_child: bool, party: bool):
        out: list[tuple[Question, dict]] = []
        errs: list[str] = []
        for prompt, raw in cells.items():
            q = qmap.get(prompt)
            if q is None:
                continue  # not an admin-defined question column
            applies = is_party_question(q) if party else person_question_applies(q, is_child=is_child)
            if not applies:
                continue
            value, err = export_import.typed_answer(q.qtype.value, q.options or [], raw)
            if err:
                errs.append(f"Row {pg.source_row}: {q.prompt}: {err}")
            elif value is not None:
                out.append((q, value))
        return out, errs

    errors: list[str] = []
    party_vals, e = resolve(pg.answers, is_child=False, party=True)
    errors += e
    comp_vals: list[list[tuple[Question, dict]]] = []
    for c in pg.companions:
        vals, e = resolve(c.answers, is_child=(c.kind == "child"), party=False)
        comp_vals.append(vals)
        errors += e
    return party_vals, comp_vals, errors


@router.post("/import", response_model=ImportResult)
async def import_guests(
    file: UploadFile = File(...),
    commit: bool = Query(False),
    ctx: WeddingCtx = Depends(editor_ctx),
    db: Session = Depends(get_db),
) -> ImportResult:
    wedding, owner = ctx.wedding, ctx.user
    require_feature(db, wedding, "import_enabled")
    """Upsert guests from a split-row XLSX (CSV also read). Dry-run by default
    (`commit=0`), returning a per-invitee preview of creates/updates/errors;
    `commit=1` applies.

    Upsert key is the `Id` column (the guest UUID), scoped to this wedding — a blank
    Id creates a new guest, an Id that doesn't resolve is an error (the `Link` column
    is display-only and never matched on). The sheet's `Tier` is authoritative when
    set; a party with more companions than the (effective) tier allows is an error,
    never silently widened. Admin-defined question columns ARE written back: when a
    row sets `Attending`, its companions + answers replace the stored RSVP."""
    data = await file.read()
    try:
        records = _read_records(file.filename, data)
    except Exception as exc:  # malformed upload
        raise HTTPException(status_code=422, detail=f"Could not read the file: {exc}") from exc

    parsed = export_import.parse_records(records)
    # Entitlement: the whole import is refused up front if the NEW invitees it
    # would create push past the plan's guest cap (dry-run included, so the
    # preview surfaces the problem before anyone commits).
    creates_planned = sum(1 for pg in parsed if not pg.guest_id)
    if creates_planned:
        check_limit(db, wedding, "max_guests", _guest_count(db, wedding), adding=creates_planned)
    qmap = {q.prompt: q for q in _questions_ordered(db, wedding)}
    phone_region = wedding_phone_region(wedding)
    results: list[ImportRowResult] = []
    created = updated = errors = 0
    people_created = people_updated = 0

    for pg in parsed:
        row_errs = list(pg.errors)

        # Normalize contacts (collect errors rather than aborting the whole import).
        email = phone = None
        try:
            email = normalize_email(pg.email)
        except ContactError as exc:
            row_errs.append(f"Row {pg.source_row}: {exc}")
        try:
            phone = normalize_phone(pg.phone, region=phone_region)
        except ContactError as exc:
            row_errs.append(f"Row {pg.source_row}: {exc}")

        # Resolve an existing guest by Id (UUID), tenant-scoped. Blank Id = create.
        guest = None
        if pg.guest_id:
            try:
                gid = UUID(pg.guest_id)
            except ValueError:
                row_errs.append(f"Row {pg.source_row}: '{pg.guest_id}' is not a valid Id")
            else:
                guest = db.execute(
                    select(Guest).where(Guest.id == gid, Guest.wedding_id == wedding.id)
                ).scalar_one_or_none()
                if guest is None:
                    row_errs.append(f"Row {pg.source_row}: Id not found in this wedding")

        # Effective tier (sheet > existing > solo) gates the companion count.
        tier = InviteTier(pg.tier) if pg.tier else None
        eff_tier = tier or (guest.invite_tier if guest else InviteTier.solo)
        caps = capabilities_for(eff_tier, wedding.content)
        adults = [c for c in pg.companions if c.kind == "adult"]
        children = [c for c in pg.companions if c.kind == "child"]
        if len(adults) > caps.max_adult_companions or len(children) > caps.max_child_companions:
            row_errs.append(
                f"Row {pg.source_row}: {len(adults)} guest / {len(children)} child "
                f"companions exceed the '{eff_tier.value}' invitation"
            )

        # Type + validate any admin-defined answers (only applied when attending).
        party_vals: list = []
        comp_vals: list = []
        if pg.attending:
            party_vals, comp_vals, answer_errs = _resolve_import_answers(pg, qmap)
            row_errs += answer_errs

        if row_errs:
            errors += 1
            results.append(
                ImportRowResult(row=pg.source_row, invitee=pg.name or pg.greeting_name or "(unnamed)", action="error", detail="; ".join(row_errs))
            )
            continue

        is_create = guest is None
        people = 1 + len(pg.companions)  # the invitee + its companions
        if not commit:
            results.append(
                ImportRowResult(
                    row=pg.source_row, invitee=pg.name or pg.greeting_name or "(unnamed)", action="create" if is_create else "update"
                )
            )
            if is_create:
                created += 1
                people_created += people
            else:
                updated += 1
                people_updated += people
            continue

        # --- Apply (commit) ---------------------------------------------------
        if guest is None:
            guest = Guest(
                wedding_id=wedding.id,
                slug=_unique_slug(db, pg.name or pg.greeting_name or ""),
                name=pg.name,
                greeting_name=pg.greeting_name,
                invite_tier=eff_tier,
                seed_meta={"source": "import"},
            )
            db.add(guest)
            created += 1
            people_created += people
        else:
            updated += 1
            people_updated += people
            guest.name = pg.name
            if tier is not None:
                guest.invite_tier = tier
        # Seed the prefill party from the sheet's Guest/Child Name rows on EVERY row
        # (regardless of Attending), so the guest's RSVP opens with names ready. Clamped
        # to the effective tier; an empty party stores NULL.
        guest.party_members = clamp_party_members(pg.companions, eff_tier, wedding.content) or None
        # Blank cells don't wipe existing values (consistent with the RSVP path).
        if email:
            guest.email = email
        if phone:
            guest.phone = phone
        if pg.greeting_name is not None:
            guest.greeting_name = pg.greeting_name
        if pg.side is not None:
            guest.side = pg.side
        if pg.relationship is not None:
            guest.relationship_label = pg.relationship
        if pg.group_name is not None:
            guest.group_name = pg.group_name
        if pg.batch is not None:
            guest.batch = pg.batch
        if pg.expected_party_size is not None:
            guest.expected_party_size = pg.expected_party_size
        guest.invited = pg.invited
        # Invite-sent flag: blank cell leaves it unchanged (never wipes).
        if pg.invite_sent is not None:
            guest.invite_sent = pg.invite_sent

        # RSVP: only touched when Attending is explicitly yes/no. When set, the
        # row's companions + answers REPLACE the stored RSVP (the sheet is the source
        # of truth for the party — same wholesale rule the companions already used).
        if pg.attending is not None:
            db.flush()  # ensure guest.id for a new guest
            rsvp = db.execute(select(Rsvp).where(Rsvp.guest_id == guest.id)).scalar_one_or_none()
            if rsvp is None:
                rsvp = Rsvp(wedding_id=wedding.id, guest_id=guest.id, attending=pg.attending)
                db.add(rsvp)
            rsvp.attending = pg.attending
            rsvp.notes = pg.notes
            if not pg.attending:
                rsvp.companions[:] = []
                rsvp.answers[:] = []
            else:
                comps = [
                    Companion(wedding_id=wedding.id, kind=CompanionKind(c.kind), name=c.name)
                    for c in pg.companions
                ]
                rsvp.companions[:] = comps
                db.flush()  # assign companion ids before attaching their answers
                new_answers = [
                    Answer(wedding_id=wedding.id, question_id=q.id, value=v, companion_id=None)
                    for q, v in party_vals
                ]
                for comp, vals in zip(comps, comp_vals):
                    new_answers += [
                        Answer(
                            wedding_id=wedding.id, question_id=q.id, value=v, companion_id=comp.id
                        )
                        for q, v in vals
                    ]
                rsvp.answers[:] = new_answers
            # Audit: an owner-driven spreadsheet import.
            stamp_rsvp(rsvp, SOURCE_IMPORT, actor=owner.email)

        results.append(
            ImportRowResult(row=pg.source_row, invitee=pg.name or pg.greeting_name or "(unnamed)", action="create" if is_create else "update")
        )

    if commit:
        db.commit()

    return ImportResult(
        committed=commit,
        created=created,
        updated=updated,
        people_created=people_created,
        people_updated=people_updated,
        errors=errors,
        rows=results,
    )


# =========================================================================
# Wedding lifecycle (SAAS_PLAN Phase 2) — approval + publication.
# =========================================================================
def _lifecycle(wedding: Wedding, *, auto: bool | None = None, trace: list | None = None) -> LifecycleResult:
    return LifecycleResult(
        wedding_id=wedding.id,
        status=wedding.status,
        published=wedding.published,
        auto_approved=auto,
        rule_trace=[RuleTraceEntry(**t) for t in (trace or [])],
    )


@router.post("/submit-approval", response_model=LifecycleResult)
def submit_for_approval(
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> LifecycleResult:
    """Owner submits the wedding for platform approval (`draft →
    pending_approval`). The auto-approval rules run immediately: all pass (and
    auto-approve is on) → instant `active`; otherwise it queues for manual
    review. Idempotent on an already-pending wedding (rules re-evaluate)."""
    wedding = ctx.wedding
    if wedding.status not in (WeddingStatus.DRAFT, WeddingStatus.PENDING_APPROVAL):
        raise HTTPException(
            status_code=409, detail="Only a draft wedding can be submitted for approval"
        )
    auto, trace = evaluate_auto_approval(db, wedding)
    wedding.status = WeddingStatus.ACTIVE if auto else WeddingStatus.PENDING_APPROVAL
    record(
        db, "wedding.submit", user=ctx.user, wedding=wedding,
        detail={"auto_approved": auto, "trace": trace},
    )
    if auto:
        record(db, "wedding.approve", user=None, wedding=wedding, detail={"auto": True})
    db.commit()
    send_email(
        settings, ctx.user.email,
        f"{wedding.couple_names}: " + ("approved!" if auto else "submitted for review"),
        "Your wedding was approved automatically — you can publish it now."
        if auto
        else "Your wedding is in the review queue; we'll email you when it's approved.",
    )
    return _lifecycle(wedding, auto=auto, trace=trace)


@router.post("/publish", response_model=LifecycleResult)
def set_published(
    payload: PublishUpdate,
    ctx: WeddingCtx = Depends(editor_ctx),
    db: Session = Depends(get_db),
) -> LifecycleResult:
    """Toggle publication — the switch that makes guest links live. Independent
    of approval: only an `active` wedding can be published (unpublishing is
    always allowed). Owner-only unless the owner granted publish rights to
    admins (`settings.admins_can_publish`)."""
    wedding = ctx.wedding
    if not _can_publish(ctx):
        raise HTTPException(status_code=403, detail="Only the owner can publish this wedding")
    if payload.published and wedding.status != WeddingStatus.ACTIVE:
        raise HTTPException(
            status_code=409, detail="This wedding must be approved before it can be published"
        )
    wedding.published = payload.published
    record(
        db, "wedding.publish" if payload.published else "wedding.unpublish",
        user=ctx.user, wedding=wedding,
    )
    db.commit()
    return _lifecycle(wedding)


@router.patch("/settings", response_model=dict)
def update_wedding_settings(
    payload: WeddingSettingsUpdate,
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
) -> dict:
    """Owner-only per-wedding admin settings (currently: whether co-admins may
    publish). Merged, not replaced; returns the effective settings blob."""
    wedding = ctx.wedding
    merged = dict(wedding.settings or {})
    for key, value in payload.model_dump(exclude_unset=True).items():
        if value is not None:
            merged[key] = value
    wedding.settings = merged
    record(db, "wedding.settings", user=ctx.user, wedding=wedding, detail=merged)
    db.commit()
    return merged


@router.delete("", response_model=LifecycleResult)
def archive_wedding(
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
) -> LifecycleResult:
    """Owner deletes the wedding — a SOFT delete: status → `archived`,
    unpublished, dashboard access ends. The 30-day undo window (reinstate, then
    hard purge) lives with the platform admin."""
    wedding = ctx.wedding
    wedding.status = WeddingStatus.ARCHIVED
    wedding.published = False
    wedding.archived_at = datetime.now(timezone.utc)  # starts the purge clock
    record(db, "wedding.archive", user=ctx.user, wedding=wedding)
    db.commit()
    return _lifecycle(wedding)


# =========================================================================
# Members — co-admin management (SAAS_PLAN Phase 3). Viewing is open to any
# member; every mutation is owner-only.
# =========================================================================
_INVITE_TTL = timedelta(days=7)


def _member_admin(db: Session, m: WeddingMember) -> MemberAdmin:
    email = m.invited_email
    display = ""
    if m.user_id:
        profile = db.get(Profile, m.user_id)
        if profile is not None:
            email = profile.email
            display = profile.display_name
    return MemberAdmin(
        id=m.id,
        user_id=m.user_id,
        email=email,
        display_name=display,
        role=m.role.value,
        status=m.status.value,
        invited_by=m.invited_by,
        created_at=m.created_at,
        invite_expires_at=m.invite_expires_at if m.status is MemberStatus.invited else None,
    )


def _active_owner_count(db: Session, wedding: Wedding) -> int:
    return db.execute(
        select(func.count()).select_from(WeddingMember).where(
            WeddingMember.wedding_id == wedding.id,
            WeddingMember.role == MemberRole.owner,
            WeddingMember.status == MemberStatus.active,
        )
    ).scalar_one()


def _get_owned_member(db: Session, wedding: Wedding, member_id: UUID) -> WeddingMember:
    m = db.execute(
        select(WeddingMember).where(
            WeddingMember.id == member_id, WeddingMember.wedding_id == wedding.id
        )
    ).scalar_one_or_none()
    if m is None:
        raise HTTPException(status_code=404, detail="Member not found")
    return m


@router.get("/members", response_model=list[MemberAdmin])
def list_members(
    wedding: Wedding = Depends(member_wedding), db: Session = Depends(get_db)
) -> list[MemberAdmin]:
    rows = db.execute(
        select(WeddingMember)
        .where(WeddingMember.wedding_id == wedding.id)
        .order_by(WeddingMember.created_at)
    ).scalars().all()
    return [_member_admin(db, m) for m in rows]


@router.post("/members", response_model=MemberInvited, status_code=201)
def invite_member(
    payload: MemberInviteCreate,
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> MemberInvited:
    """Invite a co-admin by email: an `invited` membership row + a single-use,
    expiring token (emailed; also returned as a copyable accept path). Accepting
    requires signing in with the SAME email. Re-inviting a revoked/expired email
    reuses its row with a fresh token."""
    wedding, user = ctx.wedding, ctx.user
    try:
        email = normalize_email(payload.email)
    except ContactError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if email is None:
        raise HTTPException(status_code=422, detail="An email address is required")
    email = email.lower()

    existing = db.execute(
        select(WeddingMember).where(
            WeddingMember.wedding_id == wedding.id, WeddingMember.invited_email == email
        )
    ).scalar_one_or_none()
    if existing is not None and existing.status is MemberStatus.active:
        raise HTTPException(status_code=409, detail="Already a member of this wedding")

    active_or_invited = db.execute(
        select(func.count()).select_from(WeddingMember).where(
            WeddingMember.wedding_id == wedding.id,
            WeddingMember.status != MemberStatus.revoked,
        )
    ).scalar_one()
    if existing is None or existing.status is MemberStatus.revoked:
        check_limit(db, wedding, "max_members", active_or_invited)

    token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    expires = datetime.now(timezone.utc) + _INVITE_TTL

    if existing is None:
        member = WeddingMember(
            wedding_id=wedding.id,
            invited_email=email,
            role=MemberRole(payload.role),
            status=MemberStatus.invited,
            invited_by=user.sub,
            invite_token_hash=token_hash,
            invite_expires_at=expires,
        )
        db.add(member)
    else:  # revoked or still-invited: refresh the invite
        member = existing
        member.role = MemberRole(payload.role)
        member.status = MemberStatus.invited
        member.user_id = None
        member.invited_by = user.sub
        member.invite_token_hash = token_hash
        member.invite_expires_at = expires
    db.flush()
    record(
        db, "member.invite", user=user, wedding=wedding,
        target_type="member", target_id=member.id, detail={"email": email, "role": payload.role},
    )
    db.commit()

    accept_path = f"/invites/accept?token={token}"
    send_email(
        settings, email,
        f"You're invited to help manage {wedding.couple_names}",
        f"Sign in with this email address, then open {accept_path} to accept. "
        f"The link expires in 7 days.",
    )
    return MemberInvited(member=_member_admin(db, member), accept_path=accept_path)


@router.patch("/members/{member_id}", response_model=MemberAdmin)
def update_member_role(
    member_id: UUID,
    payload: MemberRoleUpdate,
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
) -> MemberAdmin:
    wedding = ctx.wedding
    member = _get_owned_member(db, wedding, member_id)
    new_role = MemberRole(payload.role)
    if (
        member.role is MemberRole.owner
        and new_role is not MemberRole.owner
        and member.status is MemberStatus.active
        and _active_owner_count(db, wedding) <= 1
    ):
        raise HTTPException(status_code=409, detail="A wedding must keep at least one owner")
    member.role = new_role
    record(
        db, "member.role", user=ctx.user, wedding=wedding,
        target_type="member", target_id=member.id, detail={"role": payload.role},
    )
    db.commit()
    return _member_admin(db, member)


@router.delete("/members/{member_id}", response_model=MemberAdmin)
def revoke_member(
    member_id: UUID,
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> MemberAdmin:
    """Revoke a member (or cancel a pending invite). Access dies immediately —
    membership is checked per-request, nothing is cached."""
    wedding = ctx.wedding
    member = _get_owned_member(db, wedding, member_id)
    if (
        member.role is MemberRole.owner
        and member.status is MemberStatus.active
        and _active_owner_count(db, wedding) <= 1
    ):
        raise HTTPException(status_code=409, detail="A wedding must keep at least one owner")
    member.status = MemberStatus.revoked
    member.invite_token_hash = None
    member.invite_expires_at = None
    record(
        db, "member.revoke", user=ctx.user, wedding=wedding,
        target_type="member", target_id=member.id,
    )
    db.commit()
    if member.invited_email:
        send_email(
            settings, member.invited_email,
            f"Your access to {wedding.couple_names} was removed",
            "An owner removed your dashboard access for this wedding.",
        )
    return _member_admin(db, member)


@router.post("/members/{member_id}/transfer-ownership", response_model=MemberAdmin)
def transfer_ownership(
    member_id: UUID,
    ctx: WeddingCtx = Depends(owner_ctx),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> MemberAdmin:
    """Make `member_id` the owner; the current owner steps down to admin. The
    target must be an ACTIVE member (accept the invite first). The frontend
    double-confirms; the API is a single atomic step."""
    wedding, user = ctx.wedding, ctx.user
    target = _get_owned_member(db, wedding, member_id)
    if target.status is not MemberStatus.active:
        raise HTTPException(status_code=409, detail="That member hasn't accepted their invite yet")
    if target.user_id == user.sub:
        raise HTTPException(status_code=409, detail="You already own this wedding")
    target.role = MemberRole.owner
    actor_row = db.execute(
        select(WeddingMember).where(
            WeddingMember.wedding_id == wedding.id, WeddingMember.user_id == user.sub
        )
    ).scalar_one_or_none()
    if actor_row is not None:  # platform admins acting without membership skip this
        actor_row.role = MemberRole.admin
    record(
        db, "member.transfer_ownership", user=user, wedding=wedding,
        target_type="member", target_id=target.id,
    )
    db.commit()
    if target.invited_email:
        send_email(
            settings, target.invited_email,
            f"You now own {wedding.couple_names}",
            "Ownership of this wedding was transferred to you.",
        )
    return _member_admin(db, target)
