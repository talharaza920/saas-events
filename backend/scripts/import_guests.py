"""Import the planning spreadsheet into guest rows for a wedding.

Reads both side-sheets, collapses +1/kid placeholder rows into invite tiers
(app/guest_import.py), and inserts guests scoped to the wedding with unguessable
global slugs.

Usage (after migrations + seed_wedding):
    python -m scripts.import_guests \
        --file "../0_REFERENCE/Guest_List.xlsx" \
        --wedding-slug alex-and-sam [--reset] [--dry-run]
"""
from __future__ import annotations

import argparse
import warnings
from pathlib import Path

import openpyxl

from app.db import SessionLocal
from app.guest_import import build_guests, make_guest_slug
from app.models import Guest, Wedding

warnings.filterwarnings("ignore")  # the xlsx has pivot/data-validation noise

# Map normalized header text → our row key.
HEADER_MAP = {
    "name": "name",
    "side": "side",
    "relationship": "relationship",
    "group / notes": "group_name",
    "batch": "batch",
    "invite": "invite_raw",
    "prob. tier": "prob_tier",
    "prob. %": "prob_pct",
}


def _read_sheet_rows(ws) -> list[dict]:
    """Find the header row, then read data rows into normalized dicts."""
    header_idx = None
    col_map: dict[int, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "name" in cells and "side" in cells:
            header_idx = i
            for j, text in enumerate(cells):
                if text in HEADER_MAP:
                    col_map[j] = HEADER_MAP[text]
            break
    if header_idx is None:
        return []

    rows: list[dict] = []
    for row in list(ws.iter_rows(values_only=True))[header_idx + 1 :]:
        rec: dict = {}
        for j, key in col_map.items():
            val = row[j] if j < len(row) else None
            rec[key] = val.strip() if isinstance(val, str) else val
        if not rec.get("name"):
            continue
        rec["invited"] = str(rec.pop("invite_raw", "") or "").strip().lower() != "no"
        rows.append(rec)
    return rows


def read_guest_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        rows.extend(_read_sheet_rows(ws))
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--wedding-slug", default="alex-and-sam")
    parser.add_argument("--reset", action="store_true", help="delete existing guests first")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows = read_guest_rows(Path(args.file))
    guests, unresolved = build_guests(rows)
    print(f"Parsed {len(rows)} rows → {len(guests)} primary guests, {len(unresolved)} unresolved companions.")
    tier_counts: dict[str, int] = {}
    for g in guests:
        tier_counts[g.invite_tier.value] = tier_counts.get(g.invite_tier.value, 0) + 1
    print("Tier breakdown:", tier_counts)
    if unresolved:
        print("Unresolved (need admin review):", [u["name"] for u in unresolved][:20])

    if args.dry_run:
        print("[dry-run] no DB writes.")
        return

    db = SessionLocal()
    try:
        wedding = db.query(Wedding).filter_by(slug=args.wedding_slug).one_or_none()
        if wedding is None:
            raise SystemExit(f"Wedding '{args.wedding_slug}' not found — run seed_wedding first.")

        if args.reset:
            db.query(Guest).filter_by(wedding_id=wedding.id).delete()
            db.commit()

        existing = {g.name for g in db.query(Guest.name).filter_by(wedding_id=wedding.id)}
        used_slugs = {s for (s,) in db.query(Guest.slug)}
        added = 0
        for draft in guests:
            if draft.name in existing:
                continue
            slug = make_guest_slug(draft.name)
            while slug in used_slugs:
                slug = make_guest_slug(draft.name)
            used_slugs.add(slug)
            db.add(
                Guest(
                    wedding_id=wedding.id,
                    slug=slug,
                    name=draft.name,
                    # Greeting is mandatory; default to the first name for bulk seed.
                    greeting_name=(draft.name.split(" ")[0] if draft.name else "Guest"),
                    side=draft.side,
                    relationship_label=draft.relationship,
                    group_name=draft.group_name,
                    batch=draft.batch,
                    invite_tier=draft.invite_tier,
                    invited=draft.invited,
                    seed_meta={
                        **draft.seed_meta,
                        "adult_companions": draft.adult_companions,
                        "child_companions": draft.child_companions,
                    },
                )
            )
            added += 1
        # Stash unresolved companions on the wedding for admin cleanup.
        wedding.content = {
            **(wedding.content or {}),
            "_unresolved_companions": [u["name"] for u in unresolved],
        }
        db.commit()
        print(f"Inserted {added} guests into wedding '{wedding.slug}'.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
